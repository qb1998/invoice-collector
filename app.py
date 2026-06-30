#!/usr/bin/env python3
"""
发票收集器 - 单文件版 v2.2
新增功能：
1. 文件打包下载（ZIP）
2. 发票号去重并标注
3. 分类金额汇总
4. 相关文件相邻排列
5. 按购买方名称筛选发票（公司筛选）
6. 逐行删除单据并自动重排（发票编辑）
7. 解析邮件 INTERNALDATE，新增"邮件日期"列
运行方式: python app.py
然后浏览器打开 http://localhost:8000
"""
import os, re, json, ssl, socket, imaplib, email, hashlib, shutil, tempfile, asyncio, sys, zipfile, io, time
from pathlib import Path
from datetime import datetime, timedelta
from email.header import decode_header
from copy import copy
from typing import Optional

from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, Response, StreamingResponse
from pydantic import BaseModel, ConfigDict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = FastAPI(title="发票收集器", version="2.2.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===== Data Models =====
class MailConfig(BaseModel):
    email: str
    auth_code: str
    host: str = "imap.163.com"
    port: int = 993
    date_from: str
    date_to: str
    companies: list[str] = []  # 购买方筛选，[] 表示不过滤

class TaskStatus(BaseModel):
    model_config = ConfigDict(extra='allow')  # 允许挂载 state
    task_id: str
    status: str
    progress: str
    result: Optional[dict] = None

class RemoveItemRequest(BaseModel):
    source_file: str  # 唯一标识发票/辅助单据
    item_type: str   # 'invoice' | 'supporting'

tasks = {}

# ===== IMAP Helper =====
class StableIMAP4_SSL(imaplib.IMAP4_SSL):
    def _create_socket(self, timeout=None):
        addr_info = socket.getaddrinfo(self.host, self.port, 0, socket.SOCK_STREAM)
        context = self.ssl_context or ssl.create_default_context()
        for family, stype, proto, _, sockaddr in addr_info:
            try:
                sock = socket.socket(family, stype, proto)
                if timeout:
                    sock.settimeout(timeout)
                sock.connect(sockaddr)
                return context.wrap_socket(sock, server_hostname=self.host)
            except:
                continue
        raise ConnectionError('Failed to connect to %s:%d' % (self.host, self.port))

def decode_mime(v):
    if not v:
        return ''
    frags = []
    for part, enc in decode_header(v):
        if isinstance(part, bytes):
            try:
                frags.append(part.decode(enc or 'utf-8', errors='replace'))
            except:
                frags.append(part.decode('utf-8', errors='replace'))
        else:
            frags.append(part)
    return ''.join(frags).strip()

def parse_imap_date(date_str):
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    dt = datetime.strptime(date_str, '%Y-%m-%d')
    return '%02d-%s-%04d' % (dt.day, months[dt.month-1], dt.year)

# ===== PDF Extraction =====
def _ocr_pdf_with_vision(fpath, dpi=300):
    """用 macOS Vision framework 对 PDF 跑 OCR 兜底。
    适用于：纯图像 PDF（pdfplumber 提取不到文字）。
    返回 OCR 出的文字。
    """
    try:
        import io
        import tempfile
        import os
        # 把 PDF 第一页渲染为 PNG（高分辨率）
        with pdfplumber.open(str(fpath)) as pdf:
            if not pdf.pages:
                return ''
            img = pdf.pages[0].to_image(resolution=dpi)
            img_buf = io.BytesIO()
            img.original.save(img_buf, format='PNG')
            img_bytes = img_buf.getvalue()
        if not img_bytes:
            return ''
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
            f.write(img_bytes)
            tmp_png = f.name
        try:
            # 动态导入，避免在 Linux 平台 import 失败
            from Foundation import NSURL
            from Vision import VNRecognizeTextRequest, VNImageRequestHandler
            url = NSURL.fileURLWithPath_(tmp_png)
            handler = VNImageRequestHandler.alloc().initWithURL_options_(url, None)
            request = VNRecognizeTextRequest.alloc().init()
            request.setRecognitionLevel_(1)  # 1=accurate
            request.setUsesLanguageCorrection_(False)
            request.setRecognitionLanguages_(['zh-Hans', 'en-US'])
            ok, err = handler.performRequests_error_([request], None)
            if not ok:
                return ''
            results = request.results() or []
            lines = []
            for r in results:
                cands = r.topCandidates_(1)
                if cands and cands[0].confidence() >= 0.3:
                    lines.append(cands[0].string())
            return '\n'.join(lines)
        finally:
            try:
                os.unlink(tmp_png)
            except Exception:
                pass
    except Exception:
        return ''


def parse_pdf(fpath):
    """提取 PDF 文本。多次兜底：普通 -> layout -> OCR（仅 macOS）。"""
    text = ''
    try:
        with pdfplumber.open(str(fpath)) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    text += t + '\n'
            if len(text.strip()) < 50:
                text2 = ''
                for p in pdf.pages:
                    t = p.extract_text(layout=True, use_text_flow=True, x_tolerance=3, y_tolerance=3)
                    if t:
                        text2 += t + '\n'
                if len(text2.strip()) > len(text.strip()):
                    text = text2
    except Exception as e:
        return 'ERROR: %s' % e
    # OCR 兜底：纯图像 PDF（macOS Vision）
    if len(text.strip()) < 50:
        ocr_text = _ocr_pdf_with_vision(fpath)
        if ocr_text.strip() and len(ocr_text.strip()) > len(text.strip()):
            text = ocr_text
    return normalize_ocr_text(text)

def extract_seller(text, buyer_name=''):
    """提取销售方名称。兼容多种格式：
    - 销方名称 / 销货方名称 / 销售方名称 / 售方名称 / 销 名称
    - 名称：xxx（在 销/售 区块内）
    """
    # 1. 直接匹配 "销...名称："（兼容 销售方/销方/销货方 等前缀）
    m = re.search(r'销[售货方费]*\s*名\s*称[：:\s]*([^\n\r]{4,60})', text)
    if m:
        s = re.sub(r'\s+', '', m.group(1).strip())
        s = re.sub(r'(统一社会|纳税人识别号).*$', '', s).strip()
        if len(s) >= 4 and not any(h in s for h in ['规格型号', '单位', '数量', '单价']):
            if not buyer_name or s != buyer_name:
                return s
    # 2. 找 "销方信息" / "销售方信息" 区块内的 "名称：xxx"
    in_sell_section = False
    for line in text.split('\n'):
        if '销方' in line or '销售方' in line or '售方' in line or '销货方' in line:
            in_sell_section = True
        if in_sell_section and re.match(r'^\s*名\s*称[：:\s]', line):
            m2 = re.match(r'\s*名\s*称[：:\s]*(.+)', line)
            if m2:
                s = re.sub(r'\s+', '', m2.group(1).strip())
                s = re.sub(r'(统一社会|纳税人识别号).*$', '', s).strip()
                if len(s) >= 4 and (not buyer_name or s != buyer_name):
                    return s
    # 3. fallback: 找所有 "名称：xxx" 中的最后一个非 buyer
    all_names = []
    for m3 in re.finditer(r'名\s*称[：:\s]*([^\n\r]{4,50})', text):
        n = re.sub(r'\s+', '', m3.group(1).strip())
        n = re.sub(r'(统一社会|纳税人).*$', '', n).strip()
        if any(h in n[:6] for h in ['规格型号', '单位数量']):
            continue
        if len(n) >= 4 and (not buyer_name or n != buyer_name) and n not in all_names:
            all_names.append(n)
    for n in reversed(all_names):
        if len(n) >= 4:
            return n
    return ''

def extract_date(text):
    m = re.search(r'开票日期[：:\s]*(\d{4})年(\d{2})月(\d{2})日', text)
    if m:
        return '%s-%s-%s' % (m.group(1), m.group(2), m.group(3))
    m2 = re.search(r'开\s*票\s*日\s*期\s*[：:\s]*(\d{4})\s*年\s*(\d{2})\s*月\s*(\d{2})\s*日', text)
    if m2:
        return '%s-%s-%s' % (m2.group(1), m2.group(2), m2.group(3))
    return ''

def extract_amount(text):
    m = re.search(r'价税合计.*?[¥￥]?\s*([\d,]+\.\d{2})', text)
    if m:
        return m.group(1).replace(',', '')
    return ''

def extract_subtotal_tax(text):
    m = re.search(r'合\s*计\s*[:：]?\s*[¥￥]?\s*([\d,]+\.\d{2})\s*[¥￥]?\s*([\d,]+\.\d{2})', text)
    if m:
        return m.group(1).replace(',', ''), m.group(2).replace(',', '')
    return '', ''

def extract_buyer(text, default_buyer=''):
    """提取购买方名称。兼容多种格式：
    - 购方名称 / 购货方名称 / 购买方名称 / 购 名称
    - 名称：xxx（在没有 销/售 上下文时作为 fallback）
    - 「购买方信息」+ 「名称：xxx」区块
    """
    # 1. 直接匹配 "购...名称："（兼容 购买方/购方/购货方/购买方信息 等前缀）。
    #    关键：capture group 必须停在"销/售"或"纳税人识别号/统一社会信用代码"等关键字前，
    #         避免把销方或无关信息也吃进来。
    m = re.search(
        r'购[买货方]*\s*(?:信息\s*)?名\s*称[：:\s]+'
        r'([^销售\n\r]{4,40}?)'
        r'(?=[销售\n\r]|统一社会|纳税人|信用代码|开户行|账号|地址电话|$)',
        text
    )
    if m:
        b = m.group(1).strip()
        b = re.sub(r'\s+', '', b)
        # 去掉统一社会信用代码、纳税人识别号等尾部信息
        b = re.sub(r'(统一社会|纳税人|信用代码|开户行|账号|地址电话).*$', '', b).strip()
        if len(b) >= 4 and ('公司' in b or '个人' in b or len(b) >= 6):
            return b
    # 2. 区块式扫描：行级状态机，"购买方/购货方" → buyer section，"销售方/销方" → seller section
    in_section = None
    for line in text.split('\n'):
        if any(kw in line for kw in ['购买方', '购货方']):
            in_section = 'buyer'
        elif any(kw in line for kw in ['销售方', '销方', '售方', '销货方']):
            in_section = 'seller'
        if in_section == 'buyer':
            m2 = re.match(r'\s*名\s*称[：:\s]+(.+)', line)
            if m2:
                b = m2.group(1).strip()
                b = re.sub(r'\s+', '', b)
                b = re.sub(r'(统一社会|纳税人|信用代码|开户行|账号|地址电话).*$', '', b).strip()
                if len(b) >= 4:
                    return b
    return default_buyer

def determine_region(seller):
    region_map = {
        '深圳': ['深圳', '南山', '宝安', '福田', '罗湖', '龙华', '龙岗'],
        '广州': ['广州', '宝馔'],
        '河南': ['郑州', '正弘', '洛阳'],
        '海南': ['海南', '华彩'],
        '上海': ['上海', '万程'],
        '湖北': ['湖北', '科壹', '武汉'],
        '天津': ['天津'],
        '北京': ['北京'],
        '成都': ['成都'],
        '杭州': ['杭州'],
        '南京': ['南京'],
        '重庆': ['重庆'],
        '西安': ['西安'],
        '长沙': ['长沙'],
        '苏州': ['苏州'],
        '厦门': ['厦门'],
        '山东': ['山东', '济南', '青岛'],
        '四川': ['四川'],
        '湖南': ['湖南'],
        '广东': ['广东'],
        '福建': ['福建'],
    }
    for region, keywords in region_map.items():
        if any(kw in seller for kw in keywords):
            return region
    return ''

def determine_category(seller, text=''):
    combined = (seller + ' ' + text[:500])
    cat_rules = [
        ('餐饮', ['餐饮', '烧鸟', '灵鹿', '小麦', '欧力给', '博尼塔', '南万喜',
                  '川旺达', '鸟吟', '天纯', '宝馔', '烽源熠鑫', '饭店', '餐厅',
                  '美食', '食府', '火锅', '烧烤', '快餐', '小吃']),
        ('住宿', ['酒店', '威尼斯', '住宿', '宾馆', '旅馆', '民宿', '公寓', '希尔顿',
                  '洲际', '万豪', '香格里拉', '凯悦']),
        ('交通', ['出行', '打车', '滴滴', 'T3', '曹操', 'ETC', '高尔夫', '麻花科技',
                  '汽车租赁', '航空', '铁路', '火车', '高铁', '机票', '航班']),
        ('门票', ['世界之窗', '游览观光', '旅游服务', '景区', '乐园', '门票']),
        ('商超', ['超市', '商场', '正弘', '精华', '便利店', '购物']),
        ('其他', ['玩具', '科壹商贸', '斐乐', '服饰', '体育用品']),
    ]
    for cat, kws in cat_rules:
        if any(kw in combined for kw in kws):
            return cat
    return '其他'

def classify_file(filename, text):
    # === 1. 航空电子客票行程单（这是报销凭证，应当作发票）===
    if any(kw in text for kw in ['航空运输电子客票行程单', '航空运输客票', '电子客票行程单']):
        return ('invoice', '机票行程单')
    # === 2. 火车票/高铁票 ===
    if any(kw in text for kw in ['中国铁路', '12306', '电子发票(铁路电子客票)', '铁路电子客票', '高铁', '动车票']):
        return ('invoice', '火车票')
    # === 3. 网约车/出租车行程单 ===
    if any(kw in text for kw in [
        '滴滴出行行程单', '网约车行程单', '电子行程单（出租汽车）', '出租汽车发票', '滴滴出行',
        # 第三方网约车聚合平台（阳光/享道/如祺/哈啰/T3/曹操/首汽等）行程报销单
        '第三方网约车', '阳光出行', '享道出行', '如祺出行', '哈啰打车', '哈啰出行',
        'T3出行', '曹操出行', '首汽约车', '美团打车',
    ]):
        return ('invoice', '网约车发票')
    # === 4. 文件名兜底：含「行程单」/「行程报销单」+ 网约车关键词 ===
    if ('行程单' in filename or '行程报销单' in filename) and any(kw in filename for kw in [
        '滴滴', '出租', '打车', '网约', '高德', 'T3', '曹操', '首汽', '嘀嗒', '美团',
        '阳光', '享道', '如祺', '哈啰', '第三方',
    ]):
        return ('invoice', '网约车发票')
    # === 5. 文件名兜底：含「行程单」+ 航空关键词（机票、电子客票、航司）===
    if '行程单' in filename and any(kw in filename for kw in ['机票', '航空', '电子客票', '航司', '航班', '机票预订', '订座']):
        return ('invoice', '机票行程单')
    # === 5b. 文件名兜底：含「电子行程单」+ 订单号（航司电子发票 PDF 多为此种命名）===
    if '电子行程单' in filename and ('订单' in filename or '航' in filename or '机票' in filename):
        return ('invoice', '机票行程单')
    # === 5c. 携程/航旅纵横等 OTA 的「行程单」（Booking No. + 航班 + 旅客，**无金额**）===
    if any(kw in text for kw in ['Trip.com', 'trip.com', '携程', 'Booking No.', '航旅纵横', '航班管家']):
        if any(kw in text for kw in ['Itinerary', 'Itinerar', '行程单', '航班信息', 'Flight Information', 'Booking Information', '预订信息', '订单编号']):
            return ('invoice', '机票行程单-OTA')
    # === 6. 文件名兜底：含「行程单」+ 铁路/车次关键词 ===
    if '行程单' in filename and any(kw in filename for kw in ['火车', '高铁', '动车', '12306']):
        return ('invoice', '火车票')
    # === 7. 报销、出差安排材料（不是行程单）===
    if any(kw in filename for kw in ['课程日程', '酒店安排', '推荐航班', '出行须知', '课程']):
        return ('supporting', '出差行程材料')
    if any(kw in filename for kw in ['入住', '水单', '确认单']):
        return ('supporting', '酒店确认单')
    # === 8. 标准电子发票 ===
    has_invoice_content = bool(re.search(r'发票号码|价税合计|开票日期', text))
    has_invoice_kw = '发票' in filename or 'dzfp' in filename.lower() or 'qklfp' in filename.lower()
    if has_invoice_content or has_invoice_kw:
        return ('invoice', '发票')
    # === 9. 跳过明显非报销文件 ===
    skip_keywords = ['银行', '对账单', '流水', '签证', 'visa', 'eNoticeLetter',
                     '股票', '基金', '理财', '社保', '公积金', '工资']
    if any(kw in filename for kw in skip_keywords):
        return ('skip', '非报销文件')
    return ('unknown', '未识别')


def parse_air_itinerary(text):
    """解析航空运输电子客票行程单。
    返回 (seller, date, amount, remark) 二元组。
    字段缺失时对应位置为空字符串。
    """
    # 1. 承运人（销售方）
    m = re.search(r'承\s*运\s*人\s*[:： ]*\s*([^\n\r]{2,40})', text)
    if m:
        seller = m.group(1).strip()
        # 去掉尾部噪声（统一社会信用代码、纳税人识别号等）
        seller = re.sub(r'(统一社会|纳税人识别号|信用代码|开户行).*$', '', seller).strip()
    else:
        seller = ''
    # 2. 日期：优先用乘机日期，其次填开日期
    date_val = ''
    m = re.search(r'乘\s*机\s*日\s*期\s*[:： ]*\s*(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
    if m:
        date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    else:
        m = re.search(r'填\s*开\s*日\s*期\s*[:： ]*\s*(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
        if m:
            date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
        else:
            # 兜底：日期:2026-06-20
            m = re.search(r'日\s*期\s*[:： ]*\s*(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
            if m:
                date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    # 3. 金额：优先合计，其次金额
    amount = ''
    m = re.search(r'合\s*计\s*[:：]?\s*[¥￥]?\s*([\d,]+\.\d{2})', text)
    if m:
        amount = m.group(1).replace(',', '')
    else:
        m = re.search(r'金\s*额\s*[:：]?\s*[¥￥]?\s*([\d,]+\.\d{2})', text)
        if m:
            amount = m.group(1).replace(',', '')
        else:
            m = re.search(r'票\s*价\s*[:：]?\s*[¥￥]?\s*([\d,]+\.\d{2})', text)
            if m:
                amount = m.group(1).replace(',', '')
    # 4. 备注：航班号、始发站、目的站
    remark_parts = []
    m = re.search(r'航\s*班\s*号\s*[:： ]*\s*([A-Z0-9]{2}\s*\d{3,4})', text)
    if m:
        remark_parts.append('航班 ' + m.group(1).replace(' ', ''))
    m = re.search(r'始\s*发\s*站\s*[:： ]*\s*([^\n\r]{2,30})', text)
    if m:
        remark_parts.append(m.group(1).strip())
    m = re.search(r'目\s*的\s*站\s*[:： ]*\s*([^\n\r]{2,30})', text)
    if m:
        remark_parts.append('→' + m.group(1).strip())
    m = re.search(r'旅\s*客\s*姓\s*名\s*[:： ]*\s*([^\n\r]{1,20})', text)
    if m:
        remark_parts.append('旅客 ' + m.group(1).strip())
    remark = ' '.join(remark_parts)
    return seller, date_val, amount, remark


def parse_ota_itinerary(text, filename=''):
    """解析携程/Trip.com/航旅纵横等 OTA 行程单。
    返回 (seller, date, amount, remark)。**注意：OTA 行程单通常不含金额**，amount 返回空。
    """
    # OCR 后冒号常被识别为 ∶ (U+2236) 或 ：(U+FF1A)，统一规范化
    _COLON = r'[：:∶]'
    # 1. 销售方：优先航司，其次 OTA 平台
    seller = ''
    # 先找航司
    m = re.search(r'(?:航司|Airline)[/\s' + _COLON + r']*\n?\s*([^\n\r]{2,30}?)(?:\n|$)', text)
    if m:
        airline = m.group(1).strip()
        # 去重 "Airline XXX" 这种情况
        airline = re.sub(r'^[A-Za-z\s/]+', '', airline).strip()
        if airline:
            seller = airline
    if not seller:
        # 文件名兜底
        for kw in ['中国南方航空', '南方航空', '国航', '东航', '海航', '厦航', '深航']:
            if kw in text or kw in filename:
                seller = kw
                break
    if not seller:
        for kw in ['携程', 'Trip.com', '航旅纵横', '航班管家']:
            if kw in text or kw in filename:
                seller = kw
                break
    if not seller:
        seller = 'OTA行程单'
    # 2. 日期：第一个出发日期
    date_val = ''
    # 英文 "Departure 15:45, July 17, 2026"（冒号可能是 U+2236）
    m = re.search(r'Departure\s+\d{1,2}\D{0,2}\d{2},\s+(\w+)\s+(\d{1,2}),\s+(\d{4})', text)
    if m:
        months = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                  'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12}
        mo = months.get(m.group(1), 0)
        if mo:
            date_val = '%s-%02d-%02d' % (m.group(3), mo, int(m.group(2)))
    if not date_val:
        # 中文 "2026年7月17日15:45"
        m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', text)
        if m:
            date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    if not date_val:
        # 兜底 "2026-07-17"
        m = re.search(r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', text)
        if m:
            date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    # 3. 金额：OTA 行程单无金额，留空
    amount = ''
    # 4. 备注：旅客名、票号、订单号、航班号
    remark_parts = []
    m = re.search(r'(?:订单编号|Booking\s*No\.?)\s*(\d{10,20})', text)
    if m:
        remark_parts.append('订单 ' + m.group(1))
    m = re.search(r'(?:E-?ticket\s*No\.?|票号)\s*(\d{3}-?\d{10,15})', text)
    if m:
        remark_parts.append('票号 ' + m.group(1))
    m = re.search(r'(?:Airline\s*Booking\s*Reference|航司预订编码)\s*\n?\s*([A-Z0-9]{4,8})', text)
    if m:
        val = m.group(1)
        # 排除把旅客姓名（QINGBO/LI QINGBO）误识别为预订编码的情况
        if val not in ('QINGBO', 'QINGBO LI', 'LI QINGBO') and not re.match(r'^[A-Z]+\s+[A-Z]+$', val):
            remark_parts.append('预订编码 ' + val)
    # 航班号
    flights = re.findall(r'\b([A-Z]{2})\s*(\d{3,4})\b', text)
    seen = set()
    flight_strs = []
    for code, num in flights[:8]:
        key = code + num
        if key not in seen and code in ('CZ', 'CA', 'MU', 'HU', 'MF', 'ZH', '3U', 'GS', 'G5', 'FM', 'NS', 'BK', 'JD'):
            seen.add(key)
            flight_strs.append(code + num)
    if flight_strs:
        remark_parts.append('航班 ' + '/'.join(flight_strs[:3]))
    remark = ' '.join(remark_parts)
    return seller, date_val, amount, remark
    """解析火车票/高铁票。
    返回 (seller, date, amount, remark)。
    """
    # 1. 销售方：中国铁路
    seller = '中国铁路'
    # 2. 日期
    date_val = ''
    m = re.search(r'日\s*期\s*[:： ]*\s*(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
    if m:
        date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    else:
        m = re.search(r'(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})\s*\d{1,2}:\d{2}', text)
        if m:
            date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    # 3. 金额
    amount = ''
    m = re.search(r'[金票价]额\s*[:：]?\s*[¥￥]?\s*([\d,]+\.\d{2})', text)
    if m:
        amount = m.group(1).replace(',', '')
    else:
        m = re.search(r'[¥￥]\s*([\d,]+\.\d{2})', text)
        if m:
            amount = m.group(1).replace(',', '')
    # 4. 备注
    remark_parts = []
    m = re.search(r'车\s*次\s*[:： ]*\s*([A-Z]?\d{1,5})', text)
    if m:
        remark_parts.append('车次 ' + m.group(1).strip())
    m = re.search(r'(?:发车|出发|始发)\s*[站:]?\s*[:： ]*\s*([^\n\r]{2,15})', text)
    if m:
        remark_parts.append(m.group(1).strip())
    m = re.search(r'(?:到站|到达|终到)\s*[站:]?\s*[:： ]*\s*([^\n\r]{2,15})', text)
    if m:
        remark_parts.append('→' + m.group(1).strip())
    m = re.search(r'旅\s*客\s*[:： ]*\s*([^\n\r]{1,15})', text)
    if m:
        remark_parts.append('旅客 ' + m.group(1).strip())
    remark = ' '.join(remark_parts)
    return seller, date_val, amount, remark


def parse_didi_receipt(text):
    """解析网约车/出租车行程单。
    返回 (seller, date, amount, remark)。
    支持：滴滴/高德/T3/曹操/首汽/嘀嗒/美团/阳光/享道/如祺/哈啰 等。
    也支持"第三方网约车服务提供方XX—行程单"格式（聚合平台行程报销单）。
    """
    # 1. 销售方：先识别聚合方，再识别具体平台
    seller = ''
    # 聚合方/平台名
    platform_keywords = [
        '滴滴出行', '嘀嗒出行', '高德打车', '高德', 'T3出行', '曹操出行', '首汽约车', '美团打车',
        '阳光出行', '享道出行', '如祺出行', '哈啰打车', '哈啰出行', '哈啰',
    ]
    for kw in platform_keywords:
        if kw in text:
            seller = kw
            break
    # 如果是"第三方网约车服务提供方XX—行程单"格式，从标题里抠平台名
    if not seller:
        m = re.search(r'第三方网约车服务提供方([^\n\r—\-]{2,20})', text)
        if m:
            seller = '聚合网约车-' + m.group(1).strip().rstrip('—').strip()
    if not seller:
        seller = '网约车'
    # 2. 日期：优先"行程起止日期"里的"起"日期，备选"申请日期"，最后兜底第一个 4-2-2 日期
    date_val = ''
    m = re.search(r'行程起止日期\s*[:： ]*\s*(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
    if m:
        date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    else:
        m = re.search(r'(?:行程|乘车|服务|用\w*)\s*日\s*期\s*[:： ]*\s*(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
        if m:
            date_val = '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
        else:
            # 跳过"申请日期"取最早一个 4-2-2 日期
            date_candidates = re.findall(r'(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text)
            for y, mo, d in date_candidates:
                # 申请日期通常是最近日期，跳过
                if '申请' in text.split(f'{y}-{int(mo):02d}-{int(d):02d}')[0][-20:]:
                    continue
                date_val = '%s-%02d-%02d' % (y, int(mo), int(d))
                break
            if not date_val and date_candidates:
                y, mo, d = date_candidates[0]
                date_val = '%s-%02d-%02d' % (y, int(mo), int(d))
    # 3. 金额
    amount = ''
    m = re.search(r'(?:实\w*付|合\w*计|金\w*额|总\w*额)\s*[:：]?\s*[¥￥]?\s*([\d,]+\.\d{2})', text)
    if m:
        amount = m.group(1).replace(',', '')
    else:
        m = re.search(r'[¥￥]\s*([\d,]+\.\d{2})', text)
        if m:
            amount = m.group(1).replace(',', '')
    # 4. 备注：行程起止区间、笔数、起点终点
    remark_parts = []
    m = re.search(r'行程起止日期\s*[:： ]*(\d{4}[.\-/年]\d{1,2}[.\-/月]\d{1,2}\s*至\s*\d{4}[.\-/年]\d{1,2}[.\-/月]\d{1,2})', text)
    if m:
        remark_parts.append(m.group(1).replace('年', '-').replace('月', '-').replace('日', '').replace('/', '-').replace('.', '-'))
    m = re.search(r'共\s*(\d+)\s*笔行程', text)
    if m:
        remark_parts.append(m.group(1) + '笔')
    m = re.search(r'起\s*点\s*[:： ]*\s*([^\n\r]{2,30})', text)
    if m:
        remark_parts.append(m.group(1).strip())
    m = re.search(r'终\s*点\s*[:： ]*\s*([^\n\r]{2,30})', text)
    if m:
        remark_parts.append('→' + m.group(1).strip())
    remark = ' '.join(remark_parts)
    return seller, date_val, amount, remark


# OCR 常见字符替换映射（Kangxi Radicals 等容易出错）
_OCR_CHAR_MAP = {
    '⽉': '月',  # U+2F49
    '⽇': '日',  # U+2F47
    '⼀': '一',  # U+2F00
    '⽤': '用',  # U+2F64
    '⽅': '方',  # U+2F5D
    '⾏': '行',  # U+2F0F
    '⻓': '长',  # U+2EE8
    '⻘': '青',  # U+2ED8
    '⻔': '门',  # U+2ED4
    '⻩': '黄',  # U+2EE9
    '⽆': '无',
    '⼈': '人',
    '⼆': '二',
    '∶': ':',  # U+2236 ratio 冒号
}


def normalize_ocr_text(text):
    """规范化 OCR 后的常见错字（Kangxi Radicals、半角/全角符号）。"""
    if not text:
        return text
    _map = {
        '⽉': '月', '⽇': '日', '⼀': '一', '⽤': '用', '⽅': '方',
        '⾏': '行', '⻓': '长', '⻘': '青', '⻔': '门', '⻩': '黄',
        '⽆': '无', '⼈': '人', '⼆': '二', '∶': ':',
        '：': ':',  # U+FF1A fullwidth 冒号
        '，': ',',  # U+FF0C fullwidth 逗号
        '（': '(',  # U+FF08 fullwidth 左括号
        '）': ')',  # U+FF09 fullwidth 右括号
    }
    for src, dst in _map.items():
        if src in text:
            text = text.replace(src, dst)
    return text


def extract_uid_from_filename(filename):
    """Extract IMAP UID from saved filename (format: {uid}_{index}_{name})."""
    parts = filename.split('_', 2)
    if len(parts) >= 2 and parts[0].isdigit():
        return parts[0]
    return ''

# ===== Core Collection Logic =====
def collect_invoices(task_id, config):
    work_dir = Path(tempfile.mkdtemp(prefix='invoice_%s_' % task_id))
    raw_dir = work_dir / 'raw'
    raw_dir.mkdir(parents=True, exist_ok=True)
    try:
        tasks[task_id].status = 'running'
        tasks[task_id].progress = '正在连接邮箱...'
        imap = StableIMAP4_SSL(config.host, config.port)
        imap.login(config.email, config.auth_code)
        if '163.com' in config.host:
            imap.xatom('ID', '("name" "invoice-collector" "version" "1.0")')
        tasks[task_id].progress = '邮箱连接成功，正在搜索邮件...'
        imap.select('INBOX', readonly=True)
        since_date = parse_imap_date(config.date_from)
        before_date_dt = datetime.strptime(config.date_to, '%Y-%m-%d') + timedelta(days=1)
        months_list = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        before_date = '%02d-%s-%04d' % (before_date_dt.day, months_list[before_date_dt.month-1], before_date_dt.year)
        search_criteria = '(SINCE %s BEFORE %s)' % (since_date, before_date)
        status, data = imap.uid('search', None, search_criteria)
        if status != 'OK':
            raise Exception('搜索邮件失败: %s' % status)
        uids = data[0].split()
        total_emails = len(uids)
        if total_emails == 0:
            tasks[task_id].status = 'completed'
            tasks[task_id].progress = '未找到邮件'
            tasks[task_id].result = {
                'total_emails': 0, 'invoices': [], 'supporting_docs': [],
                'all_items': [], 'total_amount': 0, 'category_summary': {},
                'no_attach_invoice_emails': [],
                'skipped_emails': [], 'failed_pdfs': [], 'unknown_files': []
            }
            imap.logout()
            return
        all_attachments = []
        no_attach_invoice_emails = []
        skipped_emails = []  # 抓取/解析失败的邮件
        failed_pdfs = []  # PDF 解析失败（被记录，不丢弃原始文件）
        unknown_files = []  # 无法识别的文件（降级为待确认辅助单据）
        for i, uid_bytes in enumerate(uids):
            uid = uid_bytes.decode()
            tasks[task_id].progress = '正在处理邮件 (%d/%d)...' % (i+1, total_emails)
            subject = ''
            sender = ''
            email_date = ''
            try:
                # 一次性获取 INTERNALDATE（邮件接收时间）和邮件正文
                fetch_data = None
                fetch_error = ''
                # 重试 2 次：避免瞬时网络抖动漏抓
                for attempt in range(3):
                    try:
                        status, fd = imap.uid('fetch', uid, '(INTERNALDATE BODY.PEEK[])')
                        if status == 'OK' and fd and fd[0]:
                            fetch_data = fd
                            break
                        fetch_error = 'fetch 返回非 OK: %s' % status
                    except Exception as _e:
                        fetch_error = 'fetch 异常: %s' % str(_e)[:200]
                    if attempt < 2:
                        time.sleep(1.0 * (attempt + 1))
                if fetch_data is None:
                    skipped_emails.append({
                        'uid': uid, 'subject': '(无法获取)', 'email_date': '',
                        'reason': fetch_error or 'fetch 失败'
                    })
                    continue
                meta_raw = fetch_data[0][0]
                # 解析 INTERNALDATE -> email_date (YYYY-MM-DD)
                im = re.search(rb'INTERNALDATE "([^"]+)"', meta_raw)
                if im:
                    try:
                        ds = im.group(1).decode()
                        # 格式形如 " 8-Jun-2026 10:30:45 +0800"
                        date_part = ds.strip().split(' ')[0]
                        dt_obj = datetime.strptime(date_part, '%d-%b-%Y')
                        email_date = dt_obj.strftime('%Y-%m-%d')
                    except Exception:
                        pass
                try:
                    msg = email.message_from_bytes(fetch_data[0][1])
                except Exception as _e:
                    skipped_emails.append({
                        'uid': uid, 'subject': '(无法解析 MIME)', 'email_date': email_date,
                        'reason': 'MIME 解析失败: %s' % str(_e)[:120]
                    })
                    continue
                subject = decode_mime(msg.get('Subject', '')) if msg else ''
                sender = decode_mime(msg.get('From', '')) if msg else ''
                attachments = []
                html_body = ''
                text_body = ''
                if not msg:
                    skipped_emails.append({
                        'uid': uid, 'subject': subject, 'email_date': email_date,
                        'reason': '邮件内容为空'
                    })
                    continue
                for part in msg.walk():
                    if part.is_multipart():
                        continue
                    ct = part.get_content_type()
                    cd = part.get('Content-Disposition', '')
                    fn = part.get_filename()
                    try:
                        payload = part.get_payload(decode=True)
                        if not payload:
                            continue
                    except Exception:
                        continue
                    # 处理附件：包括 explicit attachment / inline with filename / 纯 PDF
                    is_attachment = (
                        'attachment' in cd.lower()
                        or (fn is not None and fn != '')
                        or ct == 'application/pdf'
                        or ct.startswith('image/') and ('发票' in fn or 'invoice' in fn.lower() if fn else False)
                    )
                    if is_attachment:
                        decoded_fn = decode_mime(fn) if fn else 'attachment_%s.%s' % (uid, ct.split('/')[-1] or 'bin')
                        safe_name = re.sub(r'[^\w\u4e00-\u9fff.\-]', '_', decoded_fn)
                        if not safe_name or safe_name == '_' or safe_name.startswith('.'):
                            safe_name = 'file_%s_%02d' % (uid, len(attachments))
                        out_name = '%s_%02d_%s' % (uid, len(attachments), safe_name)
                        out_path = raw_dir / out_name
                        counter = 1
                        while out_path.exists():
                            base, ext = out_name.rsplit('.', 1) if '.' in out_name else (out_name, '')
                            out_path = raw_dir / ('%s_%d.%s' % (base, counter, ext)) if ext else raw_dir / ('%s_%d' % (base, counter))
                            counter += 1
                        out_path.write_bytes(payload)
                        attachments.append({
                            'original_filename': decoded_fn,
                            'saved_filename': out_name,
                            'content_type': ct,
                            'size': len(payload),
                        })
                    elif ct == 'text/html':
                        try:
                            html_body = payload.decode('utf-8', errors='replace')
                        except Exception:
                            pass
                    elif ct == 'text/plain':
                        try:
                            text_body = payload.decode('utf-8', errors='replace')
                        except Exception:
                            pass
                urls = re.findall(r'https?://[^\s"\'<>]+', html_body)
                invoice_url_keywords = [
                    'fapiao', 'invoice', 'download', 'pdf', 'fp.', 'dppt', 'bwjf',
                    'chinatax', 'fpjx', 'e-invoice', 'qrcode', 'ewm', 'dzfp',
                    'meituan', 'ele.me', 'dianping', 'ctrip', 'fliggy', 'qunar',
                    '12306', 'alipay', 'weixin', 'qq.com', 'tencent',
                ]
                invoice_urls = []
                for u in urls:
                    u_clean = u.rstrip('>"\')')
                    if any(k in u_clean.lower() for k in invoice_url_keywords):
                        invoice_urls.append(u_clean)
                if attachments:
                    all_attachments.extend([dict(a, uid=uid, subject=subject, email_date=email_date) for a in attachments])
                else:
                    combined = (subject + text_body + html_body).lower()
                    has_invoice_kw = any(kw in combined for kw in
                        ['发票', '电子票', '行程', '报销', '凭证', '酒店', '机票', 'invoice',
                         '行程单', '水单', '入住', '打车', '出行', '美团', '饿了么',
                         '携程', '飞猪', '去哪儿', '同程', '航空', '高铁', '火车'])
                    if has_invoice_kw or invoice_urls:
                        no_attach_invoice_emails.append({
                            'uid': uid, 'subject': subject, 'from': sender,
                            'email_date': email_date,
                            'invoice_urls': invoice_urls[:10],
                        })
            except Exception as _e:
                # 顶层兜底：单封邮件出错不导致整个任务失败，但要记下来
                import traceback as _tb
                skipped_emails.append({
                    'uid': uid, 'subject': subject or '(未知)', 'email_date': email_date,
                    'reason': '处理异常: %s' % str(_e)[:120],
                    'traceback': _tb.format_exc()[:500]
                })
        imap.logout()
        tasks[task_id].progress = '已下载 %d 个附件，正在解析发票...' % len(all_attachments)
        invoices = []
        supporting_docs = []
        seen_inv_nos = {}  # inv_no -> first invoice dict
        # 建立 uid -> email_date 映射，用于在解析附件时获取邮件接收时间
        uid_to_email_date = {}
        for a in all_attachments:
            u = a.get('uid', '')
            d = a.get('email_date', '')
            if u and d and u not in uid_to_email_date:
                uid_to_email_date[u] = d
        for f in sorted(raw_dir.glob('*')):
            fname = f.name
            fpath = str(f)
            file_uid = extract_uid_from_filename(fname)
            _file_email_date = uid_to_email_date.get(file_uid, '')
            if f.suffix.lower() == '.pdf':
                text = parse_pdf(fpath)
                if text.startswith('ERROR'):
                    # PDF 解析失败：记录到 failed_pdfs，但保留 raw 文件供用户下载查看
                    failed_pdfs.append({
                        'filename': fname,
                        'error': text[6:].strip()[:200],  # 去掉 'ERROR:' 前缀
                        'email_date': _file_email_date,
                        'file_uid': file_uid,
                    })
                    # 仍然放进 supporting_docs，用户可下载查看
                    supporting_docs.append({
                        'source_file': fname, 'source_path': fpath,
                        'type': 'PDF解析失败', 'row_type': '待确认',
                        'file_uid': file_uid, 'text_snippet': '',
                        'email_date': _file_email_date,
                    })
                    continue
                file_type, subtype = classify_file(fname, text)
                if file_type == 'skip':
                    continue
                elif file_type == 'supporting':
                    supporting_docs.append({
                        'source_file': fname, 'source_path': fpath,
                        'type': subtype, 'row_type': '辅助单据',
                        'file_uid': file_uid, 'text_snippet': text[:800],
                        'email_date': _file_email_date,
                    })
                    continue
                elif file_type == 'unknown':
                    # 不再丢弃：保存为"未识别"待确认辅助单据，让用户可下载查看
                    unknown_files.append({
                        'filename': fname,
                        'reason': 'PDF 内容无法识别为发票或辅助单据',
                        'email_date': _file_email_date,
                        'file_uid': file_uid,
                    })
                    supporting_docs.append({
                        'source_file': fname, 'source_path': fpath,
                        'type': '未识别-待确认', 'row_type': '待确认',
                        'file_uid': file_uid, 'text_snippet': text[:800],
                        'email_date': _file_email_date,
                    })
                    continue
                # 行程单/火车票/网约车：用专门解析器
                _special_remark = ''
                if subtype == '机票行程单':
                    sp_seller, sp_date, sp_amount, sp_remark = parse_air_itinerary(text)
                elif subtype == '机票行程单-OTA':
                    sp_seller, sp_date, sp_amount, sp_remark = parse_ota_itinerary(text, fname)
                elif subtype == '火车票':
                    sp_seller, sp_date, sp_amount, sp_remark = parse_train_ticket(text)
                elif subtype == '网约车发票':
                    sp_seller, sp_date, sp_amount, sp_remark = parse_didi_receipt(text)
                else:
                    sp_seller = sp_date = sp_amount = sp_remark = ''
                inv_no_m = re.search(r'发票号码[：:\s]*(\d+)', text)
                if not inv_no_m:
                    inv_no_m = re.search(r'(\d{20})', text)
                if not inv_no_m and subtype == '机票行程单':
                    # 行程单用印刷序号代替发票号
                    m = re.search(r'NO[\.：:\s]*\s*([A-Z0-9]{6,15})', text, re.IGNORECASE)
                    if m:
                        inv_no = m.group(1)
                    else:
                        inv_no = ''
                elif not inv_no_m and subtype == '机票行程单-OTA':
                    # OTA 行程单用订单号作 inv_no（可能有票号）
                    m = re.search(r'(?:Booking\s*No\.?|订单编号)\s*(\d{10,20})', text)
                    if m:
                        inv_no = 'OTA-' + m.group(1)
                    else:
                        inv_no = ''
                else:
                    inv_no = inv_no_m.group(1) if inv_no_m else ''
                is_duplicate = False
                duplicate_of_inv = None
                if inv_no and inv_no in seen_inv_nos:
                    is_duplicate = True
                    duplicate_of_inv = seen_inv_nos[inv_no]
                buyer = extract_buyer(text)
                seller = extract_seller(text, buyer)
                date_val = extract_date(text)
                amount = extract_amount(text)
                no_tax, tax = extract_subtotal_tax(text)
                # 行程单/火车票/网约车：优先用专门解析器的结果
                if sp_seller:
                    seller = sp_seller
                if sp_date:
                    date_val = sp_date
                if sp_amount:
                    amount = sp_amount
                if sp_remark:
                    _special_remark = sp_remark
                if not seller:
                    fm = re.search(r'_([^_]+(?:有限公司|个体工商户|店|院))', fname)
                    if fm:
                        seller = fm.group(1)
                if not amount:
                    fa = re.search(r'(\d+\.\d{2})元?', fname)
                    if fa:
                        amount = fa.group(1)
                if not date_val:
                    fd = re.search(r'(\d{4}).(\d{2}).(\d{2})', fname)
                    if fd:
                        date_val = '%s-%s-%s' % (fd.group(1), fd.group(2), fd.group(3))
                # 若关键字段缺失，认为未识别（归为"字段缺失-待确认"）
                # 判定条件：金额为空 OR (购买方和销售方都为空)
                # 注：inv_no 不参与判定（行程单 OCR 能匹配到印刷编号但其它字段都空）
                if not amount or (not buyer and not seller):
                    missing = []
                    if not amount: missing.append('金额')
                    if not buyer: missing.append('购买方')
                    if not seller: missing.append('销售方')
                    unknown_files.append({
                        'filename': fname,
                        'reason': 'PDF 解析后关键字段缺失（%s）' % '/'.join(missing),
                        'email_date': _file_email_date,
                        'file_uid': file_uid,
                    })
                    supporting_docs.append({
                        'source_file': fname, 'source_path': fpath,
                        'type': '字段缺失-待确认', 'row_type': '待确认',
                        'file_uid': file_uid, 'text_snippet': text[:800],
                        'email_date': _file_email_date,
                    })
                    continue
                category = determine_category(seller, text)
                region = determine_region(seller)
                # 行程单/火车票/网约车：分类强制为「交通」
                if subtype in ('机票行程单', '火车票', '网约车发票'):
                    category = '交通'
                inv = {
                    'invoice_no': inv_no, 'date': date_val, 'amount': amount,
                    'amount_no_tax': no_tax, 'tax': tax, 'buyer': buyer,
                    'seller': seller, 'region': region, 'category': category,
                    'subtype': subtype,  # 保留子类型
                    'source_file': fname, 'source_path': fpath,
                    'row_type': '发票', 'file_uid': file_uid,
                    'is_duplicate': is_duplicate, 'duplicate_of': duplicate_of_inv,
                    'email_date': _file_email_date,
                    'text_snippet': text[:800],
                    'remark': _special_remark,  # 行程单/车票的备注
                }
                invoices.append(inv)
                if inv_no and not is_duplicate:
                    seen_inv_nos[inv_no] = inv
            elif f.suffix.lower() in ['.png', '.jpg', '.jpeg']:
                if any(kw in fname for kw in ['发票', '二维码', '截图']):
                    supporting_docs.append({
                        'source_file': fname, 'source_path': fpath,
                        'type': '发票截图', 'row_type': '待确认',
                        'file_uid': file_uid, 'text_snippet': '',
                        'email_date': _file_email_date,
                    })
        # Sort invoices by date+amount
        invoices.sort(key=lambda x: (x.get('date', '9999-99-99'),
                                     float(x.get('amount', '0') or 0)))
        # === Company filter (按购买方名称筛选) ===
        company_filter = set()
        if config.companies:
            for line in config.companies:
                for c in re.split(r'[,\n，、;；\s]+', line):
                    c = c.strip()
                    if c:
                        company_filter.add(c)
        all_companies = sorted({inv['buyer'] for inv in invoices if inv.get('buyer')})
        if company_filter:
            before = len(invoices)
            invoices = [inv for inv in invoices if any(c in inv.get('buyer', '') for c in company_filter)]
            tasks[task_id].progress = '按购买方筛选 %d -> %d 张发票' % (before, len(invoices))
        # === Save state for later edit operations (delete item, etc.) ===
        tasks[task_id].state = {
            'work_dir': str(work_dir),
            'raw_dir': str(raw_dir),
            'invoices': invoices,
            'supporting_docs': supporting_docs,
            'all_attachments': all_attachments,
            'all_companies': all_companies,
            'company_filter': list(company_filter),
            'email': config.email,
            'date_from': config.date_from,
            'date_to': config.date_to,
            'search_criteria': search_criteria,
            'total_emails': total_emails,
            'processed_uids': [u.decode() for u in uids],  # 已处理邮件 UID 列表（含失败但尝试过）
            'no_attach_invoice_emails': no_attach_invoice_emails,
            'skipped_emails': skipped_emails,
            'failed_pdfs': failed_pdfs,
            'unknown_files': unknown_files,
        }
        # 更新进度，让用户看到完整统计
        warn = []
        if skipped_emails: warn.append('跳过 %d 封邮件' % len(skipped_emails))
        if failed_pdfs: warn.append('%d 个 PDF 解析失败' % len(failed_pdfs))
        if unknown_files: warn.append('%d 个文件未识别' % len(unknown_files))
        warn_suffix = (' | ⚠️ ' + '，'.join(warn)) if warn else ''
        tasks[task_id].progress = '已下载 %d 个附件，已识别 %d 张发票%s' % (
            len(all_attachments), len(invoices), warn_suffix
        )
        # === Generate Excel/ZIP/result via reusable function ===
        rebuild_outputs(task_id)
        return
    except Exception as e:
        tasks[task_id].status = 'failed'
        tasks[task_id].progress = '错误: %s' % str(e)
        import traceback
        traceback.print_exc()

def rebuild_outputs(task_id):
    """从 tasks[task_id].state 读取已解析的发票/辅助单据，重新生成 Excel/ZIP/result。
    可在 collect_invoices 完成后调用，也可在删除单据后调用。"""
    state = tasks[task_id].state
    invoices = state['invoices']
    supporting_docs = state['supporting_docs']
    work_dir = Path(state['work_dir'])
    raw_dir = Path(state['raw_dir'])
    date_from = state['date_from']
    date_to = state['date_to']
    search_criteria = state['search_criteria']
    total_emails = state['total_emails']
    no_attach_invoice_emails = state['no_attach_invoice_emails']
    all_attachments = state['all_attachments']
    try:

            # === Interleave: place supporting docs next to matching invoices ===
            ordered_items = []
            used_supporting = set()
            for inv in invoices:
                ordered_items.append(('invoice', inv))
                inv_uid = inv.get('file_uid', '')
                inv_seller = inv.get('seller', '')
                inv_cat = inv.get('category', '')
                for j, sd in enumerate(supporting_docs):
                    if j in used_supporting:
                        continue
                    sd_uid = sd.get('file_uid', '')
                    sd_type = sd.get('type', '')
                    matched = False
                    # Rule 1: same email UID
                    if sd_uid and inv_uid and sd_uid == inv_uid:
                        matched = True
                    # Rule 2: hotel confirmation matches hotel invoice by seller name
                    elif sd_type == '酒店确认单' and inv_cat == '住宿':
                        sd_text = sd.get('text_snippet', '')
                        if inv_seller:
                            seller_keywords = [inv_seller[i:i+3] for i in range(0, min(len(inv_seller), 8), 2)]
                            if any(kw in sd_text for kw in seller_keywords if len(kw) >= 2):
                                matched = True
                            else:
                                hotel_invs = [i for i in invoices if i.get('category') == '住宿']
                                hotel_sds = [(k, s) for k, s in enumerate(supporting_docs)
                                             if s.get('type') == '酒店确认单' and k not in used_supporting]
                                if len(hotel_invs) == 1 and len(hotel_sds) == 1:
                                    matched = True
                    # Rule 3: itinerary matches transportation invoice
                    elif sd_type == '机票行程单' and inv_cat == '交通':
                        matched = True
                    # Rule 4: travel materials match any invoice from same date range
                    elif sd_type == '出差行程材料':
                        if inv_uid and sd_uid and inv_uid == sd_uid:
                            matched = True
                    if matched:
                        ordered_items.append(('supporting', sd))
                        used_supporting.add(j)
            # Add remaining unmatched supporting docs at the end
            for j, sd in enumerate(supporting_docs):
                if j not in used_supporting:
                    ordered_items.append(('supporting', sd))
            # === Number items: skip duplicates ===
            # 重复发票不分配序号(设为空)，其它项按出现顺序编号
            non_dup_seq = 0
            for i, (item_type, item) in enumerate(ordered_items):
                if item_type == 'invoice' and item.get('is_duplicate'):
                    item['seq_num'] = ''  # 重复发票不编号
                else:
                    non_dup_seq += 1
                    item['seq_num'] = '%03d' % non_dup_seq
            # Update duplicate references with seq numbers
            for inv in invoices:
                if inv.get('is_duplicate') and inv.get('duplicate_of'):
                    inv['duplicate_of_seq'] = inv['duplicate_of'].get('seq_num', '')
                else:
                    inv['duplicate_of_seq'] = ''
            # === Compute category summary (excluding duplicates) ===
            category_summary = {}
            for inv in invoices:
                if inv.get('is_duplicate'):
                    continue
                cat = inv.get('category', '其他')
                try:
                    amt = float(inv.get('amount', 0) or 0)
                except:
                    amt = 0
                category_summary[cat] = round(category_summary.get(cat, 0) + amt, 2)
            duplicate_count = sum(1 for inv in invoices if inv.get('is_duplicate'))
            # === Generate Excel ===
            tasks[task_id].progress = '正在生成Excel表格...'
            wb = Workbook()
            ws = wb.active
            ws.title = '发票信息'
            headers = ['序号','文件类型','地区','类目','邮件日期','发票日期','发票号码',
                       '不含税金额','税额','金额','购买方名称','销售方名称',
                       '对应发票序号','原文件名','新文件名','关联方式','备注']
            header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            data_font = Font(name='微软雅黑', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            dup_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            supp_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            inv_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
            all_rows = []
            for item_type, item in ordered_items:
                if item_type == 'invoice':
                    inv = item
                    ext = Path(inv['source_file']).suffix.lower()
                    cat = inv.get('category', '')
                    dt = inv.get('date', '')
                    amt = inv.get('amount', '')
                    is_dup = inv.get('is_duplicate', False)
                    seq = inv.get('seq_num', '')
                    if is_dup:
                        # 重复发票不重命名、不产生新文件名
                        nfn = ''
                    else:
                        nfn = '%s-%s-%s-%s%s' % (seq, cat, dt, amt, ext) if dt and amt else '%s-%s%s' % (seq, cat, ext)
                    amt_val = amt
                    try:
                        amt_val = float(amt)
                    except:
                        pass
                    remark = ''
                    if is_dup:
                        remark = '⚠️与序号%s重复，本条不计入' % inv.get('duplicate_of_seq', '')
                    # 对应发票序号：非重复发票填自身序号；重复发票填原发票序号
                    corr_seq = inv.get('duplicate_of_seq', '') if is_dup else seq
                    row_data = [
                        seq, '发票', inv['region'], cat, inv.get('email_date', ''), dt,
                        inv['invoice_no'], inv.get('amount_no_tax', ''),
                        inv.get('tax', ''), amt_val, inv['buyer'], inv['seller'],
                        corr_seq, inv['source_file'], nfn, '', remark
                    ]
                    all_rows.append({'nfn': nfn, 'data': row_data, 'type': 'invoice',
                                     'source': inv, 'is_duplicate': is_dup})
                else:
                    sd = item
                    ext = Path(sd['source_file']).suffix.lower()
                    cat = sd.get('type', '')
                    nfn = '%s-%s%s' % (sd['seq_num'], cat, ext)
                    related_seq = ''
                    sd_uid = sd.get('file_uid', '')
                    for item_type2, item2 in ordered_items:
                        if item_type2 == 'invoice':
                            i_uid = item2.get('file_uid', '')
                            if sd_uid and i_uid and sd_uid == i_uid:
                                related_seq = item2['seq_num']
                                break
                    remark = ''
                    relation = ''
                    if related_seq:
                        remark = '关联发票序号%s' % related_seq
                        relation = '关联'
                    row_data = [
                        sd['seq_num'], sd.get('row_type', '辅助单据'), '', '', sd.get('email_date', ''), '',
                        '', '', '', '', '', '',
                        related_seq, sd['source_file'], nfn, relation, remark
                    ]
                    all_rows.append({'nfn': nfn, 'data': row_data, 'type': 'supporting', 'source': sd})
            for i, r in enumerate(all_rows):
                row = i + 2
                for col, val in enumerate(r['data'], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = data_font
                    cell.alignment = center_alignment if col in [1,3,4,5,12,15] else data_alignment
                    cell.border = thin_border
                    if r.get('is_duplicate'):
                        cell.fill = dup_fill
                    elif r['type'] == 'invoice':
                        cell.fill = inv_fill
                    elif r['type'] == 'supporting':
                        cell.fill = supp_fill
            col_widths = [6, 10, 8, 8, 12, 12, 24, 12, 10, 12, 22, 30, 10, 30, 30, 8, 24]
            for i, w in enumerate(col_widths):
                ws.column_dimensions[chr(65 + i)].width = w
            # === Add category summary sheet ===
            ws2 = wb.create_sheet('分类汇总')
            ws2.cell(row=1, column=1, value='类别').font = header_font
            ws2.cell(row=1, column=2, value='金额(¥)').font = header_font
            ws2.cell(row=1, column=3, value='占比').font = header_font
            ws2.cell(row=1, column=4, value='发票数').font = header_font
            for c in range(1, 5):
                ws2.cell(row=1, column=c).fill = header_fill
                ws2.cell(row=1, column=c).alignment = header_alignment
                ws2.cell(row=1, column=c).border = thin_border
            total_cat = sum(category_summary.values())
            inv_count_by_cat = {}
            for inv in invoices:
                if inv.get('is_duplicate'):
                    continue
                cat = inv.get('category', '其他')
                inv_count_by_cat[cat] = inv_count_by_cat.get(cat, 0) + 1
            for i, (cat, amt) in enumerate(sorted(category_summary.items(), key=lambda x: -x[1])):
                r = i + 2
                ws2.cell(row=r, column=1, value=cat).border = thin_border
                ws2.cell(row=r, column=2, value=amt).border = thin_border
                ws2.cell(row=r, column=2).number_format = '#,##0.00'
                pct = '%.1f%%' % (amt / total_cat * 100) if total_cat > 0 else '0%'
                ws2.cell(row=r, column=3, value=pct).border = thin_border
                ws2.cell(row=r, column=4, value=inv_count_by_cat.get(cat, 0)).border = thin_border
            total_row = len(category_summary) + 2
            ws2.cell(row=total_row, column=1, value='合计').font = Font(name='微软雅黑', bold=True, size=11)
            ws2.cell(row=total_row, column=2, value=total_cat).font = Font(name='微软雅黑', bold=True, size=11)
            ws2.cell(row=total_row, column=2).number_format = '#,##0.00'
            ws2.cell(row=total_row, column=3, value='100%').font = Font(name='微软雅黑', bold=True, size=11)
            ws2.cell(row=total_row, column=4, value=sum(inv_count_by_cat.values())).font = Font(name='微软雅黑', bold=True, size=11)
            for c in range(1, 5):
                ws2.cell(row=total_row, column=c).border = thin_border
            ws2.column_dimensions['A'].width = 16
            ws2.column_dimensions['B'].width = 16
            ws2.column_dimensions['C'].width = 10
            ws2.column_dimensions['D'].width = 10
            xlsx_path = work_dir / '发票信息统计.xlsx'
            wb.save(str(xlsx_path))
            # Copy files to processed dir with new names (skip duplicate invoices)
            processed_dir = work_dir / 'processed'
            # Clean processed dir first to avoid stale files
            if processed_dir.exists():
                shutil.rmtree(processed_dir)
            processed_dir.mkdir(parents=True, exist_ok=True)
            for r in all_rows:
                new_name = r.get('nfn', '')
                if not new_name:
                    # 重复发票没有新文件名，不复制
                    continue
                src_name = r['data'][13]
                src_path = raw_dir / src_name if src_name else None
                if src_path and src_path.exists():
                    dst = processed_dir / new_name
                    c = 1
                    while dst.exists():
                        base, ext = new_name.rsplit('.', 1) if '.' in new_name else (new_name, '')
                        dst = processed_dir / ('%s_%d.%s' % (base, c, ext)) if ext else processed_dir / ('%s_%d' % (base, c))
                        c += 1
                    shutil.copy2(str(src_path), str(dst))
            # Build result
            total_amount = sum(float(inv.get('amount', 0) or 0) for inv in invoices if not inv.get('is_duplicate'))
            # Build all_items for frontend display (interleaved)
            all_items = []
            for item_type, item in ordered_items:
                if item_type == 'invoice':
                    inv = item
                    remark = ''
                    if inv.get('is_duplicate'):
                        remark = '⚠️重复发票，与序号%s重复' % inv.get('duplicate_of_seq', '')
                    all_items.append({
                        'seq': inv['seq_num'],
                        'item_type': 'invoice',
                        'is_duplicate': inv.get('is_duplicate', False),
                        'duplicate_of_seq': inv.get('duplicate_of_seq', ''),
                        'region': inv['region'],
                        'category': inv['category'],
                        'date': inv['date'],
                        'email_date': inv.get('email_date', ''),
                        'invoice_no': inv['invoice_no'],
                        'amount': inv['amount'],
                        'amount_no_tax': inv.get('amount_no_tax', ''),
                        'tax': inv.get('tax', ''),
                        'buyer': inv['buyer'],
                        'seller': inv['seller'],
                        'remark': remark,
                        'source_file': inv['source_file'],
                    })
                else:
                    sd = item
                    related_seq = ''
                    sd_uid = sd.get('file_uid', '')
                    for item_type2, item2 in ordered_items:
                        if item_type2 == 'invoice':
                            i_uid = item2.get('file_uid', '')
                            if sd_uid and i_uid and sd_uid == i_uid:
                                related_seq = item2['seq_num']
                                break
                    remark = ''
                    if related_seq:
                        remark = '关联发票 #%s' % related_seq
                    all_items.append({
                        'seq': sd['seq_num'],
                        'item_type': 'supporting',
                        'is_duplicate': False,
                        'duplicate_of_seq': '',
                        'region': '',
                        'category': sd.get('type', ''),
                        'date': '',
                        'email_date': sd.get('email_date', ''),
                        'invoice_no': '',
                        'amount': '',
                        'amount_no_tax': '',
                        'tax': '',
                        'buyer': '',
                        'seller': '',
                        'remark': remark,
                        'source_file': sd['source_file'],
                    })
            result = {
                'task_id': task_id,
                'search_info': {
                    'date_from': date_from,
                    'date_to': date_to,
                    'imap_search': search_criteria,
                    'emails_found': total_emails,
                    'note': '日期范围筛选的是邮件接收时间（INTERNALDATE），不是发票上的开票日期。如有 2/3 月份的发票出现在结果中，说明这些邮件是 6 月份收到的。',
                },
                'total_emails': total_emails,
                'total_attachments': len(all_attachments),
                'invoice_count': len(invoices),
                'duplicate_count': duplicate_count,
                'supporting_count': len(supporting_docs),
                'total_amount': round(total_amount, 2),
                'category_summary': category_summary,
                'all_items': all_items,
                'invoices': [
                    {
                        'seq': inv['seq_num'], 'invoice_no': inv['invoice_no'],
                        'date': inv['date'], 'amount': inv['amount'],
                        'amount_no_tax': inv.get('amount_no_tax', ''),
                        'tax': inv.get('tax', ''),
                        'buyer': inv['buyer'], 'seller': inv['seller'],
                        'region': inv['region'], 'category': inv['category'],
                        'is_duplicate': inv.get('is_duplicate', False),
                        'duplicate_of_seq': inv.get('duplicate_of_seq', ''),
                        'remark': '⚠️重复发票，与序号%s重复' % inv.get('duplicate_of_seq', '') if inv.get('is_duplicate') else '',
                    } for inv in invoices
                ],
                'supporting_docs': [
                    {
                        'seq': sd['seq_num'], 'type': sd.get('type', ''),
                        'source_file': sd['source_file'],
                    } for sd in supporting_docs
                ],
                'no_attach_invoice_emails': [
                    {
                        'uid': e['uid'], 'subject': e['subject'],
                        'from': e['from'], 'invoice_urls': e.get('invoice_urls', []),
                    } for e in no_attach_invoice_emails[:20]
                ],
                'skipped_emails': [
                    {
                        'uid': e.get('uid', ''),
                        'subject': e.get('subject', ''),
                        'email_date': e.get('email_date', ''),
                        'reason': e.get('reason', ''),
                    } for e in state.get('skipped_emails', [])
                ],
                'failed_pdfs': [
                    {
                        'filename': p.get('filename', ''),
                        'error': p.get('error', ''),
                        'email_date': p.get('email_date', ''),
                    } for p in state.get('failed_pdfs', [])
                ],
                'unknown_files': [
                    {
                        'filename': u.get('filename', ''),
                        'reason': u.get('reason', ''),
                        'email_date': u.get('email_date', ''),
                    } for u in state.get('unknown_files', [])
                ],
                'work_dir': str(work_dir),
            }

            tasks[task_id].status = 'completed'
            duplicate_count = sum(1 for inv in invoices if inv.get('is_duplicate'))
            total_amount = sum(float(inv.get('amount', 0) or 0) for inv in invoices if not inv.get('is_duplicate'))
            tasks[task_id].progress = '完成！共 %d 张发票（含 %d 张重复），%d 个辅助文件，合计 ¥%.2f' % (
                len(invoices), duplicate_count, len(supporting_docs), total_amount)
            tasks[task_id].result = result
    except Exception as e:
        tasks[task_id].status = 'failed'
        tasks[task_id].progress = '错误: %s' % str(e)
        import traceback
        traceback.print_exc()

# ===== API Endpoints =====
@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "2.2.0"}

@app.post("/api/collect")
async def start_collection(config: MailConfig, background_tasks: BackgroundTasks):
    task_id = hashlib.md5(
        ("%s%s%s%s" % (config.email, config.date_from, config.date_to, datetime.now().isoformat())).encode()
    ).hexdigest()[:12]
    tasks[task_id] = TaskStatus(
        task_id=task_id, status='pending',
        progress='任务已创建，等待执行...', result=None
    )
    background_tasks.add_task(collect_invoices, task_id, config)
    return {"task_id": task_id, "message": "任务已创建"}

@app.get("/api/status/{task_id}")
async def get_status(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    return tasks[task_id]

@app.get("/api/download/{task_id}")
async def download_excel(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = tasks[task_id]
    if task.status != 'completed' or not task.result:
        raise HTTPException(status_code=400, detail="任务未完成或无结果")
    work_dir = Path(task.result['work_dir'])
    xlsx_path = work_dir / '发票信息统计.xlsx'
    if not xlsx_path.exists():
        raise HTTPException(status_code=404, detail="Excel文件不存在")
    return FileResponse(
        str(xlsx_path),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename='发票信息统计_%s.xlsx' % task_id
    )

@app.get("/api/download-all/{task_id}")
async def download_all_files(task_id: str):
    """Download all files as a ZIP archive, including the Excel summary.
    Note: Duplicate invoices are EXCLUDED from the ZIP (matches Excel table content)."""
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = tasks[task_id]
    if task.status != 'completed' or not task.result:
        raise HTTPException(status_code=400, detail="任务未完成或无结果")
    work_dir = Path(task.result['work_dir'])
    processed_dir = work_dir / 'processed'
    if not processed_dir.exists():
        raise HTTPException(status_code=404, detail="文件目录不存在")
    zip_buffer = io.BytesIO()
    file_count = 0
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        xlsx_path = work_dir / '发票信息统计.xlsx'
        if xlsx_path.exists():
            zf.write(str(xlsx_path), '发票信息统计.xlsx')
            file_count += 1
        for f in sorted(processed_dir.glob('*')):
            if f.is_file():
                zf.write(str(f), f.name)
                file_count += 1
    zip_buffer.seek(0)
    from urllib.parse import quote
    zip_filename = '发票文件包_%s.zip' % task_id
    encoded_name = quote(zip_filename)
    return StreamingResponse(
        zip_buffer,
        media_type='application/zip',
        headers={'Content-Disposition': "attachment; filename=\"%s\"; filename*=UTF-8''%s" % ('invoice_%s.zip' % task_id, encoded_name)}
    )

@app.get("/api/files/{task_id}")
async def list_files(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = tasks[task_id]
    if task.status != 'completed' or not task.result:
        raise HTTPException(status_code=400, detail="任务未完成")
    work_dir = Path(task.result['work_dir'])
    processed_dir = work_dir / 'processed'
    if not processed_dir.exists():
        return {"files": []}
    files = []
    for f in sorted(processed_dir.glob('*')):
        files.append({'name': f.name, 'size': f.stat().st_size})
    return {"files": files}

@app.get("/api/file/{task_id}/{filename:path}")
async def download_file(task_id: str, filename: str, inline: int = 0):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = tasks[task_id]
    # 支持未完成的任务预览（用户在结果页展开文件预览）
    if not hasattr(task, 'state') or task.state is None:
        raise HTTPException(status_code=400, detail="任务状态已丢失")
    state = task.state
    work_dir = Path(state['work_dir'])
    # 安全检查：filename 不能包含 ..
    if '..' in filename or filename.startswith('/'):
        raise HTTPException(status_code=400, detail="非法文件名")
    file_path = work_dir / 'processed' / filename
    if not file_path.exists():
        file_path = work_dir / 'raw' / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="文件不存在")
    # inline=1 时浏览器内联预览（PDF 用浏览器内置查看器）
    if inline:
        return FileResponse(str(file_path), filename=filename, content_disposition_type='inline')
    return FileResponse(str(file_path), filename=filename)

@app.post("/api/task/{task_id}/remove-item")
async def remove_item(task_id: str, req: RemoveItemRequest):
    """从任务结果中删除指定单据（发票或辅助单据），重新生成 Excel/ZIP/result。
    source_file: 要删除的原始文件名
    item_type: 'invoice' | 'supporting'
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = tasks[task_id]
    if not hasattr(task, 'state') or task.state is None:
        raise HTTPException(status_code=400, detail="任务状态已丢失，无法编辑")
    state = task.state
    invoices = state.get('invoices', [])
    supporting_docs = state.get('supporting_docs', [])
    removed = False
    removed_source = ''
    if req.item_type == 'invoice':
        new_invoices = []
        for inv in invoices:
            if inv.get('source_file') == req.source_file:
                removed = True
                removed_source = inv.get('source_file', '')
                # 顺便清理引用了此发票的重复发票
                for inv2 in invoices:
                    if inv2.get('is_duplicate') and inv2.get('duplicate_of') is inv:
                        inv2['_to_delete'] = True
                continue
            new_invoices.append(inv)
        # 把引用了已删除发票的重复发票也删掉
        final_invoices = [inv for inv in new_invoices if not inv.get('_to_delete')]
        if len(final_invoices) != len(new_invoices):
            removed = True
        state['invoices'] = final_invoices
    elif req.item_type == 'supporting':
        new_sds = [sd for sd in supporting_docs if sd.get('source_file') != req.source_file]
        if len(new_sds) != len(supporting_docs):
            removed = True
            removed_source = req.source_file
        state['supporting_docs'] = new_sds
    else:
        raise HTTPException(status_code=400, detail="item_type 必须为 invoice 或 supporting")
    if not removed:
        raise HTTPException(status_code=404, detail="未找到要删除的单据")
    # 重新生成 Excel/ZIP/result
    rebuild_outputs(task_id)
    return {
        'success': True,
        'removed_source': removed_source,
        'result': task.result,
    }

@app.get("/api/task/{task_id}/companies")
async def get_companies(task_id: str):
    """获取任务涉及的所有购买方名称（用于公司筛选下拉/选择）。"""
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    task = tasks[task_id]
    if not hasattr(task, 'state') or task.state is None:
        return {"companies": []}
    return {"companies": task.state.get('all_companies', [])}

# ===== Embedded Frontend =====
HTML_CONTENT = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>发票收集器 - 智能发票管理平台</title>
<style>
:root{--primary:#4F46E5;--primary-hover:#4338CA;--primary-light:#EEF2FF;--success:#059669;--success-light:#D1FAE5;--warning:#D97706;--warning-light:#FEF3C7;--danger:#DC2626;--danger-light:#FEE2E2;--gray-50:#F9FAFB;--gray-100:#F3F4F6;--gray-200:#E5E7EB;--gray-300:#D1D5DB;--gray-400:#9CA3AF;--gray-500:#6B7280;--gray-600:#4B5563;--gray-700:#374151;--gray-800:#1F2937;--gray-900:#111827;--shadow-lg:0 10px 15px -3px rgba(0,0,0,0.1),0 4px 6px -4px rgba(0,0,0,0.1);--radius:12px;--radius-sm:8px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Hiragino Sans GB','Microsoft YaHei',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;color:var(--gray-800)}
.container{max-width:1100px;margin:0 auto;padding:24px 16px}
.header{text-align:center;padding:40px 0 32px;color:#fff}
.header h1{font-size:36px;font-weight:700;letter-spacing:-.5px;margin-bottom:8px}
.header p{font-size:16px;opacity:.85}
.card{background:#fff;border-radius:var(--radius);box-shadow:var(--shadow-lg);padding:32px;margin-bottom:24px;transition:transform .2s}
.card:hover{transform:translateY(-1px)}
.card-title{font-size:20px;font-weight:600;color:var(--gray-800);margin-bottom:24px;display:flex;align-items:center;gap:10px}
.card-title .icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:18px}
.icon-blue{background:var(--primary-light);color:var(--primary)}
.icon-green{background:var(--success-light);color:var(--success)}
.icon-orange{background:var(--warning-light);color:var(--warning)}
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}
@media(max-width:640px){.form-grid{grid-template-columns:1fr}}
.form-group{display:flex;flex-direction:column;gap:6px}
.form-group.full-width{grid-column:1/-1}
.form-label{font-size:14px;font-weight:500;color:var(--gray-600)}
.form-input,.form-select{padding:12px 16px;border:1.5px solid var(--gray-200);border-radius:var(--radius-sm);font-size:15px;color:var(--gray-800);transition:border-color .2s,box-shadow .2s;outline:0;background:#fff}
.form-input:focus,.form-select:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(79,70,229,.1)}
.form-input::placeholder{color:var(--gray-400)}
.form-hint{font-size:12px;color:var(--gray-400)}
.btn{padding:14px 32px;border:0;border-radius:var(--radius-sm);font-size:16px;font-weight:600;cursor:pointer;transition:all .2s;display:inline-flex;align-items:center;gap:8px}
.btn-primary{background:var(--primary);color:#fff}
.btn-primary:hover{background:var(--primary-hover);transform:translateY(-1px)}
.btn-primary:disabled{background:var(--gray-300);cursor:not-allowed;transform:none}
.btn-success{background:var(--success);color:#fff}
.btn-success:hover{opacity:.9;transform:translateY(-1px)}
.btn-outline{background:#fff;color:var(--primary);border:1.5px solid var(--primary)}
.btn-outline:hover{background:var(--primary-light)}
.btn-mini{display:inline-flex;align-items:center;justify-content:center;padding:4px 10px;font-size:13px;font-weight:600;color:#fff;background:var(--primary);border-radius:6px;text-decoration:none;transition:all .2s}
.btn-mini:hover{background:var(--primary-hover);transform:translateY(-1px)}
.btn-row-delete{background:transparent;border:1px solid #FCA5A5;color:#DC2626;width:30px;height:30px;border-radius:6px;cursor:pointer;font-size:14px;transition:all .2s;display:inline-flex;align-items:center;justify-content:center}
.btn-row-delete:hover{background:#FEE2E2;border-color:#DC2626;transform:scale(1.05)}
.progress-bar-container{background:var(--gray-100);border-radius:99px;height:10px;overflow:hidden;margin-bottom:12px}
.progress-bar{height:100%;background:linear-gradient(90deg,var(--primary),#818CF8);border-radius:99px;transition:width .5s ease;width:0}
.progress-text{font-size:14px;color:var(--gray-600);text-align:center}
.spinner{display:inline-block;width:36px;height:36px;border:3px solid rgba(79,70,229,.2);border-top-color:var(--primary);border-radius:50%;animation:spin .8s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.stats-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:16px;margin-bottom:24px}
@media(max-width:640px){.stats-grid{grid-template-columns:repeat(2,1fr)}}
.stat-card{background:#fff;border-radius:var(--radius-sm);padding:20px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.1)}
.stat-value{font-size:28px;font-weight:700;color:var(--primary)}
.stat-value.red{color:#DC2626}
.stat-value.orange{color:#D97706}
.stat-label{font-size:13px;color:var(--gray-500);margin-top:4px}
.category-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px}
.cat-card{background:var(--gray-50);border-radius:var(--radius-sm);padding:16px;text-align:center;border:1.5px solid var(--gray-200);transition:all .2s}
.cat-card:hover{border-color:var(--primary);transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,0,0,.08)}
.cat-name{font-size:13px;color:var(--gray-500);margin-bottom:6px}
.cat-amount{font-size:20px;font-weight:700;color:var(--gray-800)}
.cat-pct{font-size:11px;color:var(--gray-400);margin-top:4px}
.cat-count{font-size:11px;color:var(--gray-400);margin-top:2px}
.table-container{overflow-x:auto;border-radius:var(--radius-sm);border:1px solid var(--gray-200)}
table{width:100%;border-collapse:collapse;font-size:14px}
thead th{background:var(--primary);color:#fff;padding:12px 10px;text-align:left;font-weight:500;white-space:nowrap;position:sticky;top:0}
tbody td{padding:10px;border-bottom:1px solid var(--gray-100);white-space:nowrap}
tbody tr:hover{background:var(--primary-light)}
tbody tr.invoice-row{background:#F0FDF4}
tbody tr.supporting-row{background:#FEF3C7}
tbody tr.duplicate-row{background:#FEE2E2 !important}
tbody tr.duplicate-row:hover{background:#FECACA !important}
.badge{display:inline-block;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:500}
.badge-invoice{background:#D1FAE5;color:#059669}
.badge-supporting{background:#FEF3C7;color:#D97706}
.badge-duplicate{background:#FEE2E2;color:#DC2626}
.action-row{display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap}
.email-alert{background:var(--warning-light);border:1px solid #F59E0B;border-radius:var(--radius-sm);padding:16px;margin-bottom:16px}
.email-alert h4{color:var(--warning);margin-bottom:8px}
.email-alert ul{list-style:disc;padding-left:20px;font-size:14px;color:var(--gray-700)}
.email-alert li{margin-bottom:4px}
.email-alert a{color:var(--primary);text-decoration:underline}
.skipped-alert{background:#FEE2E2;border:1.5px solid #DC2626;border-radius:var(--radius-sm);padding:14px 16px;margin-bottom:12px}
.skipped-alert h4{color:#991B1B;margin-bottom:6px;display:flex;align-items:center;gap:6px}
.skipped-alert .alert-body{font-size:13px;color:#7F1D1D;max-height:240px;overflow-y:auto;background:#fff;border:1px solid #FECACA;border-radius:6px;padding:8px 10px;margin-top:8px}
.skipped-alert .alert-body .row{padding:4px 0;border-bottom:1px dashed #FEE2E2}
.skipped-alert .alert-body .row:last-child{border-bottom:none}
.skipped-alert code{background:#FEE2E2;padding:1px 4px;border-radius:3px;font-size:12px;color:#7F1D1D}
.search-info-banner{background:linear-gradient(135deg,#EFF6FF 0%,#DBEAFE 100%);border:1px solid #93C5FD;border-radius:var(--radius-sm);padding:14px 18px;margin-bottom:20px;font-size:13px;color:var(--gray-700)}
.search-info-title{font-weight:600;color:#1E40AF;margin-bottom:8px;font-size:14px}
.search-info-row{margin-bottom:4px;display:flex;align-items:center;flex-wrap:wrap;gap:4px}
.search-info-label{color:var(--gray-600);font-weight:500}
.search-info-banner code{background:rgba(30,64,175,.1);color:#1E40AF;padding:2px 6px;border-radius:4px;font-family:monospace;font-size:12px}
.search-info-tip{margin-top:8px;padding:8px 10px;background:rgba(255,255,255,.6);border-radius:6px;color:#1E3A8A;line-height:1.5}
.search-info-tip strong{color:#1E40AF}
.steps{display:flex;justify-content:center;gap:8px;margin-bottom:32px}
.step{display:flex;align-items:center;gap:8px;padding:8px 16px;border-radius:99px;font-size:14px;font-weight:500;background:rgba(255,255,255,.15);color:rgba(255,255,255,.6);transition:all .3s}
.step.active{background:rgba(255,255,255,.25);color:#fff}
.step.done{background:rgba(255,255,255,.3);color:#fff}
.step-num{width:24px;height:24px;border-radius:50%;background:rgba(255,255,255,.2);display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700}
.step.active .step-num{background:#fff;color:var(--primary)}
.step.done .step-num{background:#10B981;color:#fff}
.step-arrow{color:rgba(255,255,255,.3)}
.footer{text-align:center;padding:32px 0 16px;color:rgba(255,255,255,.5);font-size:13px}
.toast{position:fixed;top:20px;right:20px;padding:14px 20px;border-radius:var(--radius-sm);color:#fff;font-weight:500;font-size:14px;z-index:9999;transform:translateX(120%);transition:transform .3s ease;max-width:360px}
.toast.show{transform:translateX(0)}
.toast-success{background:var(--success)}
.toast-error{background:var(--danger)}
.toast-info{background:var(--primary)}
.error-box{display:none;margin-top:16px;padding:16px 18px;background:#FEF2F2;border:1.5px solid #FCA5A5;border-radius:var(--radius-sm);color:#991B1B}
.error-box.show{display:block}
.error-box-title{font-weight:600;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.error-box-msg{font-size:13px;line-height:1.6;word-break:break-all}
.error-box-hint{margin-top:10px;padding-top:10px;border-top:1px dashed #FCA5A5;font-size:12px;color:#7F1D1D}
.error-box-hint b{color:#991B1B}
.security-note{background:var(--primary-light);border-radius:var(--radius-sm);padding:14px 16px;margin-top:16px;display:flex;gap:10px;align-items:flex-start;font-size:13px;color:var(--gray-600)}
.security-note .lock-icon{font-size:18px;flex-shrink:0;margin-top:1px}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
.fade-in{animation:fadeIn .4s ease}
.legend{display:flex;gap:16px;margin-bottom:16px;font-size:13px;color:var(--gray-500)}
.legend-item{display:flex;align-items:center;gap:6px}
.legend-dot{width:12px;height:12px;border-radius:3px}
.legend-dot-green{background:#D1FAE5;border:1px solid #059669}
.legend-dot-yellow{background:#FEF3C7;border:1px solid #D97706}
.legend-dot-red{background:#FEE2E2;border:1px solid #DC2626}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>🧾 发票收集器</h1>
<p>输入邮箱信息，一键收集全部发票并生成统计表格</p>
</div>
<div class="steps">
<div class="step active" id="step1"><span class="step-num">1</span>配置邮箱</div>
<span class="step-arrow">→</span>
<div class="step" id="step2"><span class="step-num">2</span>收集发票</div>
<span class="step-arrow">→</span>
<div class="step" id="step3"><span class="step-num">3</span>查看结果</div>
</div>
<div class="card" id="configCard">
<div class="card-title"><span class="icon icon-blue">📧</span>邮箱配置</div>
<form id="configForm" onsubmit="startCollection(event)">
<div class="form-grid">
<div class="form-group">
<label class="form-label">邮箱地址</label>
<input type="email" class="form-input" id="email" placeholder="yourname@163.com" required>
<span class="form-hint">支持163、QQ、Gmail等支持IMAP的邮箱</span>
</div>
<div class="form-group">
<label class="form-label">IMAP授权码</label>
<input type="password" class="form-input" id="authCode" placeholder="输入IMAP授权码" required>
<span class="form-hint">非登录密码，需在邮箱设置中开启IMAP并获取授权码</span>
</div>
<div class="form-group">
<label class="form-label">IMAP服务器</label>
<select class="form-select" id="imapHost" onchange="updatePort()">
<option value="imap.163.com">163邮箱</option>
<option value="imap.qq.com">QQ邮箱</option>
<option value="imap.gmail.com">Gmail</option>
<option value="imap.outlook.com">Outlook</option>
<option value="imap.126.com">126邮箱</option>
<option value="imap.sina.com">新浪邮箱</option>
<option value="custom">自定义服务器</option>
</select>
</div>
<div class="form-group" id="customHostGroup" style="display:none">
<label class="form-label">自定义IMAP服务器</label>
<input type="text" class="form-input" id="customHost" placeholder="imap.example.com">
</div>
<div class="form-group">
<label class="form-label">开始日期</label>
<input type="date" class="form-input" id="dateFrom" required>
</div>
<div class="form-group">
<label class="form-label">结束日期</label>
<input type="date" class="form-input" id="dateTo" required>
</div>
<div class="form-group">
<label class="form-label">按购买方筛选 <span style="font-weight:normal;color:#6B7280;font-size:12px;">（可选，逗号/换行分隔；留空表示不过滤）</span></label>
<textarea class="form-input" id="companies" rows="2" placeholder="例如：李清博（个人）&#10;郑州方信新材料"></textarea>
</div>
</div>
<div class="security-note">
<span class="lock-icon">🔒</span>
<span>您的邮箱授权码仅用于本次IMAP连接读取邮件，不会被存储或上传。所有处理均在本地完成。</span>
</div>
<div style="text-align:center;margin-top:28px;">
<button type="submit" class="btn btn-primary" id="startBtn">🚀 开始收集发票</button>
</div>
</form>
</div>
<div class="card" id="progressCard" style="display:none">
<div class="card-title"><span class="icon icon-blue">⏳</span>正在收集发票</div>
<div class="progress-bar-container"><div class="progress-bar" id="progressBar"></div></div>
<div class="progress-text" id="progressText">正在连接邮箱...</div>
<div style="text-align:center;margin-top:20px;"><div class="spinner" id="progressSpinner"></div></div>
<div class="error-box" id="errorBox">
<div class="error-box-title">⚠️ 收集失败</div>
<div class="error-box-msg" id="errorBoxMsg"></div>
<div class="error-box-hint">
<b>常见原因：</b><br>
• 授权码错误（请到邮箱设置中获取<b>IMAP授权码</b>，不是登录密码）<br>
• 邮箱未开启 IMAP 服务（需在网页版邮箱设置中开启）<br>
• 网络/防火墙拦截（公司/校园网可能屏蔽 993 端口）<br>
• 邮箱地址与 IMAP 服务器不匹配（如 163 邮箱选了 QQ 服务器）<br>
</div>
<div style="margin-top:14px;text-align:center;">
<button type="button" class="btn btn-primary" onclick="backToConfig()" style="padding:10px 24px;font-size:14px;">↩ 重新填写</button>
</div>
</div>
</div>
<div id="resultSection" style="display:none">
<div class="stats-grid fade-in">
<div class="stat-card"><div class="stat-value" id="statEmails">0</div><div class="stat-label">扫描邮件</div></div>
<div class="stat-card"><div class="stat-value" id="statInvoices">0</div><div class="stat-label">发票数量</div></div>
<div class="stat-card"><div class="stat-value orange" id="statDuplicates">0</div><div class="stat-label">重复发票</div></div>
<div class="stat-card"><div class="stat-value" id="statSupporting">0</div><div class="stat-label">辅助文件</div></div>
<div class="stat-card"><div class="stat-value red" id="statAmount">¥0</div><div class="stat-label">发票总额</div></div>
</div>
<div id="searchInfoBanner" class="search-info-banner fade-in" style="display:none">
<div class="search-info-title">🔍 搜索条件与说明</div>
<div class="search-info-row"><span class="search-info-label">邮件日期范围：</span><span id="searchInfoRange">-</span></div>
<div class="search-info-row"><span class="search-info-label">IMAP 搜索语句：</span><code id="searchInfoCiteria">-</code></div>
<div class="search-info-row"><span class="search-info-label">实际找到邮件：</span><span id="searchInfoFound">-</span></div>
<div class="search-info-tip">💡 日期范围筛选的是<strong>邮件接收时间</strong>（即 IMAP INTERNALDATE），不是发票上的开票日期。表格中"邮件日期"列才是筛选依据。如有 2/3 月份的发票出现，说明这些邮件是 6 月份收到的。</div>
</div>
<div id="categorySection" style="display:none" class="card fade-in">
<div class="card-title"><span class="icon icon-green">💰</span>分类金额汇总</div>
<div class="category-grid" id="categoryGrid"></div>
</div>
<div class="action-row fade-in">
<button class="btn btn-success" onclick="downloadExcel()">📥 下载Excel表格</button>
<button class="btn btn-primary" onclick="downloadAllFiles()">📦 下载全部文件(ZIP)</button>
<button class="btn btn-outline" onclick="newCollection()">🔄 新建收集任务</button>
</div>
<div id="emailAlertSection" style="display:none">
<div class="email-alert fade-in">
<h4>⚠️ 发现无附件的发票邮件</h4>
<p style="margin-bottom:8px;font-size:14px;">以下邮件包含发票信息但无附件，可能需要手动下载：</p>
<ul id="emailAlertList"></ul>
</div>
</div>
<div id="skippedEmailsSection" style="display:none">
<div class="skipped-alert fade-in">
<h4>🚫 跳过的邮件（<span id="skippedCount">0</span> 封）</h4>
<p style="font-size:13px;color:#7F1D1D;margin-bottom:4px;">以下邮件在抓取/解析过程中出错，未被收集。请检查邮箱设置或网络后重新收集，或联系开发排查：</p>
<div class="alert-body" id="skippedEmailsList"></div>
</div>
</div>
<div id="failedPdfsSection" style="display:none">
<div class="skipped-alert fade-in">
<h4>📄 PDF 解析失败（<span id="failedPdfsCount">0</span> 个）</h4>
<p style="font-size:13px;color:#7F1D1D;margin-bottom:4px;">以下 PDF 文件因加密、扫描件或格式问题无法提取文字，已作为「待确认」辅助单据保留在文件包中，可手动查看：</p>
<div class="alert-body" id="failedPdfsList"></div>
</div>
</div>
<div id="unknownFilesSection" style="display:none">
<div class="skipped-alert fade-in" style="background:#FEF3C7;border-color:#D97706">
<h4 style="color:#92400E;">❓ 未识别的文件（<span id="unknownFilesCount">0</span> 个）</h4>
<p style="font-size:13px;color:#78350F;margin-bottom:4px;">以下文件无法识别为发票或常见辅助单据，已作为「待确认」辅助单据保留。请人工检查是否为相关单据：</p>
<div class="alert-body" id="unknownFilesList" style="background:#fff;border-color:#FDE68A"></div>
</div>
</div>

<!-- PDF/图片预览模态框 -->
<div id="previewModal" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.7);z-index:9999;justify-content:center;align-items:center;">
<div style="background:#fff;width:90%;max-width:1100px;height:90%;border-radius:8px;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,0.4);">
<div style="padding:12px 16px;border-bottom:1px solid #E5E7EB;display:flex;justify-content:space-between;align-items:center;">
<div id="previewModalTitle" style="font-weight:600;color:#1F2937;font-size:14px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1;margin-right:12px;"></div>
<div style="display:flex;gap:8px;align-items:center;">
<a id="previewDownloadLink" href="#" target="_blank" rel="noopener" style="color:#2563EB;text-decoration:none;font-size:13px;padding:4px 8px;">⬇ 下载</a>
<button onclick="closePreview()" style="background:#F3F4F6;border:none;border-radius:4px;padding:4px 12px;cursor:pointer;font-size:14px;color:#6B7280;">✕ 关闭</button>
</div>
</div>
<div id="previewModalBody" style="flex:1;overflow:auto;background:#F9FAFB;display:flex;justify-content:center;align-items:flex-start;padding:8px;"></div>
</div>
</div>
<div class="card fade-in">
<div class="card-title"><span class="icon icon-green">📊</span>发票详情</div>
<div class="legend">
<div class="legend-item"><span class="legend-dot legend-dot-green"></span>发票</div>
<div class="legend-item"><span class="legend-dot legend-dot-yellow"></span>辅助单据</div>
<div class="legend-item"><span class="legend-dot legend-dot-red"></span>重复发票</div>
</div>
<div class="table-container">
<table>
<thead><tr><th>序号</th><th>类型</th><th>地区</th><th>类目</th><th>邮件日期</th><th>发票日期</th><th>发票号码</th><th>金额</th><th>购买方</th><th>销售方</th><th>备注</th><th>操作</th></tr></thead>
<tbody id="invoiceTableBody"></tbody>
</table>
</div>
</div>
<div class="card fade-in">
<div class="card-title"><span class="icon icon-blue">📁</span>已收集文件</div>
<div id="filesList"></div>
</div>
</div>
</div>
<div class="footer">发票收集器 v2.0 · 所有数据本地处理，安全可靠</div>
<div class="toast" id="toast"></div>
<script>
const API_BASE=window.location.origin;
let currentTaskId=null,pollInterval=null;
const today=new Date();
const threeMonthsAgo=new Date(today);
threeMonthsAgo.setMonth(threeMonthsAgo.getMonth()-3);
document.getElementById('dateFrom').value=threeMonthsAgo.toISOString().split('T')[0];
document.getElementById('dateTo').value=today.toISOString().split('T')[0];
document.getElementById('email').addEventListener('blur',function(){
const email=this.value.toLowerCase();
const select=document.getElementById('imapHost');
if(email.includes('163.com'))select.value='imap.163.com';
else if(email.includes('qq.com'))select.value='imap.qq.com';
else if(email.includes('gmail.com'))select.value='imap.gmail.com';
else if(email.includes('outlook.com')||email.includes('hotmail.com'))select.value='imap.outlook.com';
else if(email.includes('126.com'))select.value='imap.126.com';
else if(email.includes('sina.com'))select.value='imap.sina.com';
updatePort();
});
function updatePort(){
const select=document.getElementById('imapHost');
document.getElementById('customHostGroup').style.display=select.value==='custom'?'block':'none';
}
function showToast(msg,type='info'){
const toast=document.getElementById('toast');
toast.textContent=msg;
toast.className='toast toast-'+type+' show';
setTimeout(()=>toast.classList.remove('show'),4000);
}
function setStep(stepNum){
['step1','step2','step3'].forEach((id,i)=>{
const el=document.getElementById(id);
el.className='step';
if(i+1<stepNum)el.classList.add('done');
if(i+1===stepNum)el.classList.add('active');
});
}
async function startCollection(e){
e.preventDefault();
const email=document.getElementById('email').value.trim();
const authCode=document.getElementById('authCode').value.trim();
const imapHost=document.getElementById('imapHost').value==='custom'
?document.getElementById('customHost').value.trim()
:document.getElementById('imapHost').value;
const dateFrom=document.getElementById('dateFrom').value;
const dateTo=document.getElementById('dateTo').value;
const companiesRaw=document.getElementById('companies').value.trim();
// 把 textarea 文本切成多个公司名（支持换行、逗号、中文逗号、分号、空白）
const companies=companiesRaw?companiesRaw.split(/[\n,，;；、\s]+/).map(function(s){return s.trim();}).filter(Boolean):[];
if(!email||!authCode||!dateFrom||!dateTo){showToast('请填写所有必填项','error');return;}
const startBtn=document.getElementById('startBtn');
startBtn.disabled=true;
startBtn.textContent='⏳ 正在创建任务...';
document.getElementById('errorBox').classList.remove('show');
document.getElementById('configCard').style.display='none';
document.getElementById('progressCard').style.display='block';
setStep(2);
try{
const resp=await fetch(API_BASE+'/api/collect',{
method:'POST',
headers:{'Content-Type':'application/json'},
body:JSON.stringify({email,auth_code:authCode,host:imapHost,port:993,date_from:dateFrom,date_to:dateTo,companies:companies})
});
if(!resp.ok){
const errText=await resp.text().catch(()=>resp.statusText);
throw new Error('HTTP '+resp.status+': '+errText);
}
const data=await resp.json();
currentTaskId=data.task_id;
pollInterval=setInterval(pollStatus,2000);
showToast('任务已创建，正在收集...','info');
}catch(err){
showStartError('启动任务失败',err.message);
}
}
async function pollStatus(){
if(!currentTaskId)return;
try{
const resp=await fetch(API_BASE+'/api/status/'+currentTaskId);
const data=await resp.json();
const progressMap={'pending':5,'running':50,'completed':100,'failed':100};
const pct=progressMap[data.status]||0;
document.getElementById('progressBar').style.width=pct+'%';
document.getElementById('progressText').textContent=data.progress;
if(data.status==='completed'){clearInterval(pollInterval);resetStartBtn();showResults(data.result);setStep(3);}
else if(data.status==='failed'){
clearInterval(pollInterval);
document.getElementById('progressText').textContent='❌ ' + data.progress;
showErrorBox(data.progress);
showToast('收集失败，请查看下方详情','error');
}
}catch(err){console.error('Poll error:',err);}
}
function showErrorBox(errMsg){
const box=document.getElementById('errorBox');
const msg=document.getElementById('errorBoxMsg');
msg.textContent=errMsg;
box.classList.add('show');
}
function showStartError(title,detail){
resetStartBtn();
document.getElementById('progressCard').style.display='none';
document.getElementById('configCard').style.display='block';
setStep(1);
showErrorBox(title+'：'+detail);
setTimeout(()=>{document.getElementById('errorBox').classList.remove('show');},8000);
}
function resetStartBtn(){
const btn=document.getElementById('startBtn');
btn.disabled=false;
btn.textContent='🚀 开始收集发票';
}
function backToConfig(){
clearInterval(pollInterval);
currentTaskId=null;
document.getElementById('progressCard').style.display='none';
document.getElementById('configCard').style.display='block';
document.getElementById('errorBox').classList.remove('show');
setStep(1);
resetStartBtn();
}
// ====== PDF/图片预览模态框 ======
function openPreview(filename){
if(!filename||!currentTaskId){return;}
const modal=document.getElementById('previewModal');
const body=document.getElementById('previewModalBody');
const title=document.getElementById('previewModalTitle');
const dl=document.getElementById('previewDownloadLink');
title.textContent=filename;
const encoded=encodeURIComponent(filename);
const url=API_BASE+'/api/file/'+currentTaskId+'/'+encoded+'?inline=1';
const dlUrl=API_BASE+'/api/file/'+currentTaskId+'/'+encoded;
dl.href=dlUrl;
dl.setAttribute('download',filename);
// 判断文件类型
const lower=filename.toLowerCase();
body.innerHTML='<div style="padding:40px;color:#6B7280;">加载中...</div>';
if(lower.endsWith('.pdf')){
// 浏览器内置 PDF 查看器
body.innerHTML='<iframe src="'+url+'" style="width:100%;height:100%;border:0;background:#fff;" title="'+filename.replace(/"/g,'&quot;')+'"></iframe>';
}else if(lower.match(/\.(png|jpg|jpeg|gif|webp|bmp)$/)){
body.innerHTML='<img src="'+url+'" style="max-width:100%;max-height:100%;object-fit:contain;" alt="'+filename.replace(/"/g,'&quot;')+'" />';
}else{
// 其它文件类型：显示下载链接
body.innerHTML='<div style="padding:40px;text-align:center;"><p style="color:#6B7280;margin-bottom:16px;">该文件类型不支持在线预览，请下载查看：</p><a href="'+dlUrl+'" target="_blank" rel="noopener" download="'+filename.replace(/"/g,'&quot;')+'" style="display:inline-block;padding:8px 20px;background:#2563EB;color:#fff;border-radius:4px;text-decoration:none;">⬇ 下载 '+filename+'</a></div>';
}
modal.style.display='flex';
// ESC 键关闭
const escHandler=function(e){if(e.key==='Escape'){closePreview();document.removeEventListener('keydown',escHandler);}};
document.addEventListener('keydown',escHandler);
// 点击背景关闭
modal.onclick=function(e){if(e.target===modal){closePreview();}};
}
function closePreview(){
const modal=document.getElementById('previewModal');
modal.style.display='none';
modal.onclick=null;
const body=document.getElementById('previewModalBody');
body.innerHTML='';
}
function showResults(result){
document.getElementById('progressCard').style.display='none';
document.getElementById('resultSection').style.display='block';
document.getElementById('statEmails').textContent=result.total_emails;
document.getElementById('statInvoices').textContent=result.invoice_count;
document.getElementById('statDuplicates').textContent=result.duplicate_count||0;
document.getElementById('statSupporting').textContent=result.supporting_count;
document.getElementById('statAmount').textContent='¥'+result.total_amount.toLocaleString('zh-CN',{minimumFractionDigits:2});
// 显示搜索条件信息
if(result.search_info){
document.getElementById('searchInfoBanner').style.display='block';
document.getElementById('searchInfoRange').textContent=result.search_info.date_from+' ~ '+result.search_info.date_to;
document.getElementById('searchInfoCiteria').textContent=result.search_info.imap_search;
document.getElementById('searchInfoFound').textContent=result.search_info.emails_found+' 封';
}
// Render category summary
if(result.category_summary&&Object.keys(result.category_summary).length>0){
document.getElementById('categorySection').style.display='block';
const grid=document.getElementById('categoryGrid');
const total=result.total_amount||0;
let html='';
const sorted=Object.entries(result.category_summary).sort((a,b)=>b[1]-a[1]);
sorted.forEach(function(item){
const cat=item[0],amt=item[1];
const pct=total>0?(amt/total*100).toFixed(1):'0';
const count=result.invoices?result.invoices.filter(function(inv){return inv.category===cat&&!inv.is_duplicate;}).length:0;
html+='<div class="cat-card"><div class="cat-name">'+cat+'</div><div class="cat-amount">¥'+amt.toLocaleString('zh-CN',{minimumFractionDigits:2})+'</div><div class="cat-pct">'+pct+'%</div><div class="cat-count">'+count+'张</div></div>';
});
grid.innerHTML=html;
}
// Render table using all_items (interleaved order)
const tbody=document.getElementById('invoiceTableBody');
tbody.innerHTML='';
if(result.all_items&&result.all_items.length>0){
result.all_items.forEach(function(item){
const tr=document.createElement('tr');
if(item.is_duplicate){
tr.className='duplicate-row';
}else if(item.item_type==='invoice'){
tr.className='invoice-row';
}else{
tr.className='supporting-row';
}
let badge='';
if(item.is_duplicate){
badge='<span class="badge badge-duplicate">重复</span>';
}else if(item.item_type==='invoice'){
badge='<span class="badge badge-invoice">发票</span>';
}else{
badge='<span class="badge badge-supporting">辅助</span>';
}
let amountCell='-';
if(item.amount){
amountCell='<span style="color:#DC2626;font-weight:600;">¥'+item.amount+'</span>';
}
tr.innerHTML='<td style="'+(item.is_duplicate?'color:#9CA3AF;font-style:italic;':'')+'">'+(item.seq||'—')+'</td><td>'+badge+'</td><td>'+(item.region||'-')+'</td><td>'+(item.category||'-')+'</td><td style="font-size:12px;color:#6B7280;">'+(item.email_date||'-')+'</td><td>'+(item.date||'-')+'</td><td style="font-family:monospace;font-size:12px;">'+(item.invoice_no||'-')+'</td><td>'+amountCell+'</td><td>'+(item.buyer||'-')+'</td><td>'+(item.seller||'-')+'</td><td style="font-size:12px;'+(item.is_duplicate?'color:#DC2626;font-weight:500;':'')+'">'+(item.remark||'')+'</td><td style="white-space:nowrap;"><button class="btn-row-preview" data-source="'+(item.source_file||'').replace(/"/g,'&quot;')+'" data-fname="'+(item.source_file||'').replace(/"/g,'&quot;')+'" title="预览文件" style="background:#DBEAFE;border:1px solid #93C5FD;color:#1E40AF;border-radius:4px;padding:2px 6px;cursor:pointer;font-size:12px;margin-right:4px;">👁</button><button class="btn-row-delete" data-source="'+(item.source_file||'').replace(/"/g,'&quot;')+'" data-type="'+item.item_type+'" title="删除此单据" style="background:#FEE2E2;border:1px solid #FCA5A5;color:#991B1B;border-radius:4px;padding:2px 6px;cursor:pointer;font-size:12px;">🗑</button></td>';
const delBtn=tr.querySelector('.btn-row-delete');
if(delBtn){delBtn.addEventListener('click',function(){removeItem(this.getAttribute('data-source')||'',this.getAttribute('data-type')||'');});}
const previewBtn=tr.querySelector('.btn-row-preview');
if(previewBtn){previewBtn.addEventListener('click',function(){openPreview(this.getAttribute('data-fname')||'');});}
tbody.appendChild(tr);
});
}else{
// Fallback: use separate invoices and supporting_docs
result.invoices.forEach(function(inv){
const tr=document.createElement('tr');
tr.className=inv.is_duplicate?'duplicate-row':'invoice-row';
let badge=inv.is_duplicate?'<span class="badge badge-duplicate">重复</span>':'<span class="badge badge-invoice">发票</span>';
const _sf=inv.source_file||'';
const _sfEsc=_sf.replace(/"/g,'&quot;');
tr.innerHTML='<td style="'+(inv.is_duplicate?'color:#9CA3AF;font-style:italic;':'')+'">'+(inv.seq||'—')+'</td><td>'+badge+'</td><td>'+inv.region+'</td><td>'+inv.category+'</td><td>'+inv.date+'</td><td style="font-family:monospace;font-size:12px;">'+inv.invoice_no+'</td><td style="color:#DC2626;font-weight:600;">¥'+inv.amount+'</td><td>'+inv.buyer+'</td><td>'+inv.seller+'</td><td style="font-size:12px;color:#DC2626;">'+(inv.remark||'')+'</td><td style="white-space:nowrap;"><button class="btn-row-preview" data-fname="'+_sfEsc+'" title="预览" style="background:#DBEAFE;border:1px solid #93C5FD;color:#1E40AF;border-radius:4px;padding:2px 6px;cursor:pointer;font-size:12px;margin-right:4px;">👁</button>-</td>';
const pbtn=tr.querySelector('.btn-row-preview');
if(pbtn){pbtn.addEventListener('click',function(){openPreview(this.getAttribute('data-fname')||'');});}
tbody.appendChild(tr);
});
result.supporting_docs.forEach(function(sd){
const tr=document.createElement('tr');
tr.className='supporting-row';
const _sf=sd.source_file||'';
const _sfEsc=_sf.replace(/"/g,'&quot;');
tr.innerHTML='<td>'+sd.seq+'</td><td><span class="badge badge-supporting">辅助</span></td><td>-</td><td>'+sd.type+'</td><td>-</td><td>-</td><td>-</td><td>-</td><td>-</td><td></td><td style="white-space:nowrap;"><button class="btn-row-preview" data-fname="'+_sfEsc+'" title="预览" style="background:#DBEAFE;border:1px solid #93C5FD;color:#1E40AF;border-radius:4px;padding:2px 6px;cursor:pointer;font-size:12px;">👁</button>-</td>';
const pbtn=tr.querySelector('.btn-row-preview');
if(pbtn){pbtn.addEventListener('click',function(){openPreview(this.getAttribute('data-fname')||'');});}
tbody.appendChild(tr);
});
}
if(result.no_attach_invoice_emails&&result.no_attach_invoice_emails.length>0){
document.getElementById('emailAlertSection').style.display='block';
const ul=document.getElementById('emailAlertList');
ul.innerHTML='';
result.no_attach_invoice_emails.forEach(function(e){
const li=document.createElement('li');
let html='<strong>'+e.subject+'</strong> (来自: '+e.from+')';
if(e.invoice_urls&&e.invoice_urls.length>0){
html+='<br>发票链接: '+e.invoice_urls.map(function(u){return '<a href="'+u+'" target="_blank" rel="noopener">下载链接</a>';}).join(' | ');
}
li.innerHTML=html;
ul.appendChild(li);
});
}
// 跳过的邮件
if(result.skipped_emails&&result.skipped_emails.length>0){
document.getElementById('skippedEmailsSection').style.display='block';
document.getElementById('skippedCount').textContent=result.skipped_emails.length;
const body=document.getElementById('skippedEmailsList');
body.innerHTML='';
result.skipped_emails.forEach(function(e){
const row=document.createElement('div');
row.className='row';
row.innerHTML='<strong>UID '+e.uid+'</strong>'
+(e.email_date?' <code>'+e.email_date+'</code>':'')
+'<br>主题: '+e.subject
+'<br>原因: <code>'+(e.reason||'未知')+'</code>';
body.appendChild(row);
});
}
// PDF 解析失败
if(result.failed_pdfs&&result.failed_pdfs.length>0){
document.getElementById('failedPdfsSection').style.display='block';
document.getElementById('failedPdfsCount').textContent=result.failed_pdfs.length;
const body=document.getElementById('failedPdfsList');
body.innerHTML='';
result.failed_pdfs.forEach(function(p){
const row=document.createElement('div');
row.className='row';
const _fnEsc=(p.filename||'').replace(/"/g,'&quot;');
row.innerHTML='<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px;">'
+'<div style="flex:1;"><strong>'+p.filename+'</strong>'
+(p.email_date?' <code>'+p.email_date+'</code>':'')
+'<br>错误: <code>'+(p.error||'未知')+'</code></div>'
+'<button class="btn-row-preview" data-fname="'+_fnEsc+'" title="预览文件" style="background:#DBEAFE;border:1px solid #93C5FD;color:#1E40AF;border-radius:4px;padding:4px 10px;cursor:pointer;font-size:12px;white-space:nowrap;">👁 预览</button>'
+'</div>';
const pbtn=row.querySelector('.btn-row-preview');
if(pbtn){pbtn.addEventListener('click',function(){openPreview(this.getAttribute('data-fname')||'');});}
body.appendChild(row);
});
}
// 未识别的文件
if(result.unknown_files&&result.unknown_files.length>0){
document.getElementById('unknownFilesSection').style.display='block';
document.getElementById('unknownFilesCount').textContent=result.unknown_files.length;
const body=document.getElementById('unknownFilesList');
body.innerHTML='';
result.unknown_files.forEach(function(u){
const row=document.createElement('div');
row.className='row';
const _fnEsc=(u.filename||'').replace(/"/g,'&quot;');
row.innerHTML='<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px;">'
+'<div style="flex:1;"><strong>'+u.filename+'</strong>'
+(u.email_date?' <code>'+u.email_date+'</code>':'')
+'<br>原因: '+(u.reason||'未知')+'</div>'
+'<button class="btn-row-preview" data-fname="'+_fnEsc+'" title="预览文件" style="background:#DBEAFE;border:1px solid #93C5FD;color:#1E40AF;border-radius:4px;padding:4px 10px;cursor:pointer;font-size:12px;white-space:nowrap;">👁 预览</button>'
+'</div>';
const pbtn=row.querySelector('.btn-row-preview');
if(pbtn){pbtn.addEventListener('click',function(){openPreview(this.getAttribute('data-fname')||'');});}
body.appendChild(row);
});
}
loadFilesList();
var msg='收集完成！共'+result.invoice_count+'张发票';
if(result.duplicate_count>0)msg+='（含'+result.duplicate_count+'张重复）';
if(result.skipped_emails&&result.skipped_emails.length>0)msg+='，⚠️ 跳过 '+result.skipped_emails.length+' 封邮件';
if(result.failed_pdfs&&result.failed_pdfs.length>0)msg+='，⚠️ '+result.failed_pdfs.length+' 个 PDF 失败';
if(result.unknown_files&&result.unknown_files.length>0)msg+='，⚠️ '+result.unknown_files.length+' 个文件未识别';
showToast(msg,'success');
}
async function loadFilesList(){
if(!currentTaskId)return;
try{
const resp=await fetch(API_BASE+'/api/files/'+currentTaskId);
const data=await resp.json();

const container=document.getElementById('filesList');
if(!data.files||data.files.length===0){container.innerHTML='<p style="color:var(--gray-400);">暂无文件</p>';return;}
let html='<div style="display:grid;grid-template-columns:1fr auto auto;gap:6px 12px;font-size:13px;align-items:center;">';
data.files.forEach(function(f){
var size=f.size>1024*1024?(f.size/1024/1024).toFixed(1)+' MB':(f.size/1024).toFixed(0)+' KB';
var encoded=encodeURIComponent(f.name);
html+='<div style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="'+f.name+'">'+f.name+'</div><div style="color:var(--gray-400);white-space:nowrap;">'+size+'</div><div><a class="btn-mini" href="'+API_BASE+'/api/file/'+currentTaskId+'/'+encoded+'" target="_blank" download title="下载该文件">⬇</a></div>';
});
html+='</div>';
container.innerHTML=html;
}catch(err){console.error('Load files error:',err);}
}
function downloadExcel(){
if(!currentTaskId)return;
window.open(API_BASE+'/api/download/'+currentTaskId,'_blank');
}
function downloadAllFiles(){
if(!currentTaskId)return;
showToast('正在打包文件，请稍候...','info');
window.open(API_BASE+'/api/download-all/'+currentTaskId,'_blank');
}
async function removeItem(sourceFile,itemType){
if(!currentTaskId){showToast('没有可编辑的任务','error');return;}
if(!sourceFile){showToast('缺少文件标识','error');return;}
const label=itemType==='invoice'?'该发票':'该辅助单据';
if(!confirm('确认从结果中删除'+label+'？\n文件名：'+sourceFile+'\n\n注：此操作会重新生成 Excel/统计表/文件包（序号会重排）。'))return;
try{
showToast('正在删除并重新生成...','info');
const resp=await fetch(API_BASE+'/api/task/'+currentTaskId+'/remove-item',{
method:'POST',
headers:{'Content-Type':'application/json'},
body:JSON.stringify({source_file:sourceFile,item_type:itemType})
});
if(!resp.ok){
const err=await resp.json().catch(function(){return {};});
throw new Error(err.detail||('请求失败: '+resp.status));
}
const data=await resp.json();
showToast('已删除，序号已重排','success');
// 用后端返回的最新 result 重新渲染
if(data.result){
showResults(data.result);
}
}catch(err){
showToast('删除失败: '+err.message,'error');
console.error(err);
}
}
function newCollection(){
currentTaskId=null;
if(pollInterval)clearInterval(pollInterval);
document.getElementById('configCard').style.display='block';
document.getElementById('progressCard').style.display='none';
document.getElementById('resultSection').style.display='none';
document.getElementById('categorySection').style.display='none';
document.getElementById('emailAlertSection').style.display='none';
document.getElementById('skippedEmailsSection').style.display='none';
document.getElementById('failedPdfsSection').style.display='none';
document.getElementById('unknownFilesSection').style.display='none';
document.getElementById('progressBar').style.width='0%';
setStep(1);
}
</script>
</body>
</html>"""

@app.get("/")
async def root():
    return HTMLResponse(content=HTML_CONTENT)

if __name__ == '__main__':
    import uvicorn
    print("=" * 50)
    print("  发票收集器 v2.0 已启动")
    print("  请在浏览器打开: http://localhost:8000")
    print("  按 Ctrl+C 停止服务")
    print("=" * 50)
    uvicorn.run(app, host='0.0.0.0', port=8000)
